import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci

SRC = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
COLS = ("C D E F G H I J K L M O P Q S T X "
        "II IJ IK IM IN IO IP IQ IR IS IT IU IV IW IX IY "
        "JF JG JH JI JJ JK JM JN JO JP JQ JR JT JU JV JW "
        "KE KF KJ KK KM KN KP KQ KR KS KT KV KW KX KY KZ LA LB LC "
        "LF LL LR NZ OL OM ON OO "
        "QD QE QF QG QH QK QL QM QN QO "
        "QR QS QT QU QV QW QX QY QZ RA RB RC "
        "SV SW SX SY TC TD TE TF TH TI TJ TK TO TP TQ TR TT TU TV TW UB UC UD "
        "UF UG UH UJ").split()
IDX = {ci(c): c for c in COLS}
MAXC = max(IDX)

WANT = set()
# Sport 520/560/600/660 first PVC + first HYP + model header rows
for r in (812,820,829,830,831,832,833,834,835,837,838,839,840,841,842,843,844,
          846,854,863,871,1533,1534,1535):
    WANT.add(r)
# scan range for Patrol rows too, add later after we see names
WANT |= set(range(880, 950))

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb['Boat Module']
out = {}
rn = 0
for row in ws.iter_rows(min_row=1, max_row=1000, max_col=MAXC, values_only=True):
    rn += 1
    if rn not in WANT: continue
    cells = {}
    for i, v in enumerate(row, start=1):
        letter = IDX.get(i)
        if letter and v not in (None, ''):
            cells[letter] = v
    if cells: out[rn] = cells
wb.close()
json.dump(out, open('nm_rows.json','w',encoding='utf-8'), indent=1, default=str)
print('rows captured', sorted(out))
