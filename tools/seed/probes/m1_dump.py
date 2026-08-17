"""Dump the Motor Library into extracts/m1_*.json. READ-ONLY on the workbook.

`Motor Module (1).xlsx` in Downloads is a TRUNCATED ZIP and raises BadZipFile;
`probes/motor_00_salvage.py` rebuilds it, and `Copy of Motor Module (1).xlsx` is
that salvaged copy. Same source MPF_GROUND_TRUTH.md used.

ROW EXTENT. This used to stop at row 345, which cut the library off in the
MIDDLE of the factory boat+engine block: the `Powerplants - Haines Signature`
banner is row 342 and the packages under it run to 584. The library's real last
populated row is 589. FITMENT_RULES.md §5.1 J7/J8 needs that block — those
values are what 291 Jeanneau and Haines motor slots name, and with the dump
stopping at 345 they resolved to nothing at all.
"""
import openpyxl, json, sys
from pathlib import Path
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Copy of Motor Module (1).xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"

KEEP = set()
for a, b in [("A", "CY"), ("EY", "FI")]:
    KEEP |= set(range(ci(a), ci(b) + 1))
MAXC = max(KEEP)

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Motor Library"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
style = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=600, min_col=1, max_col=MAXC), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        if i not in KEEP:
            continue
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = v if isinstance(v, (int, float, bool)) else str(v)
        if i == 3 and v is not None and str(v).strip() != "":
            try:
                style[r] = {"sz": c.font.sz, "b": c.font.b}
            except Exception:
                pass
    if d:
        data[r] = d

json.dump(data, open(OUT + "m1_data.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
json.dump(style, open(OUT + "m1_style.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("rows", len(data), file=sys.stderr)
wb.close()
