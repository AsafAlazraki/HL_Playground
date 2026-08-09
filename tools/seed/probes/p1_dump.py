import openpyxl, json, sys
from openpyxl.utils import get_column_letter as gl

P = r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
OUT = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Parts Maintenance"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=3722, min_col=1, max_col=30), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = v if isinstance(v, (int, float, bool)) else str(v)
    if d:
        data[r] = d
json.dump(data, open(OUT + "p1_parts.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("parts rows", len(data), file=sys.stderr)
wb.close()
