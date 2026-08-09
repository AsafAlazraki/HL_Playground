from openpyxl import load_workbook
SRC=r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
CODES={'HET005','HET008','HEP014','HEC196','HER007','HEE006','HEF 031','HET045','HEP106','HEO107','HEP018-W'}
wb=load_workbook(SRC, read_only=True, data_only=True)
ws=wb['Highfield']
rn=0
for row in ws.iter_rows(min_row=1,max_row=1398,max_col=16,values_only=True):
    rn+=1
    code=row[2]
    if isinstance(code,str) and code.strip() in CODES:
        print(rn,'| C',row[2],'| D',row[3],'| E',row[4],'| F',row[5],'| G USD',row[6],'| H AUD',row[7],'| L Trade',row[11],'| P Retail',row[15])
wb.close()
