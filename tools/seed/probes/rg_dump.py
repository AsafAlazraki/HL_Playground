"""Dump Registration Module.xlsx · sheet "Registration Costs" -> extracts/rg_rego.json.

READ-ONLY. Open it, never save it.

WHY THIS SHEET. SERVICE_AND_THEMES.md §3.1: registration is ONE concept —
"a third-party statutory charge, looked up by band from one shared table, never
marked up" — and that sentence is true of a boat and of a trailer with no edits.
The business itself says so: `Boat Transfer Fee` and `Trailer Transfer Fee` are
two rows at the SAME $32.55 with two revenue codes. It duplicated the ROW, not
the TABLE.

SHAPE, per the spec, read first-hand and re-confirmed here:
  C6            'AS at 1/7/25'   — the vintage of every fee below
  row 8         the header row (C Band · G REV Code · J CTD · K SELL)
  C8/C15/C21/C29  four SECTION banner rows in the same column as the data,
                exactly the way the Boat Module and Trailer Module draw bands.
                They are the only place the SUBJECT of a fee is written down.
  rows 9-13     Boat Registration              banded on hull LENGTH
  rows 16-19    Trailer Registration           banded on ATM MASS
  rows 22-27    Other Fees & Charges           not banded
  rows 30-33    Boat Registration - Pensioner / Concession Card Holder

The whole sheet is C3:K34, so the dump is C..K over rows 1..40 — small enough
that taking all of it costs nothing and guessing a narrower window costs a
column.
"""
import openpyxl, json, sys
from pathlib import Path
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Registration Module.xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Registration Costs"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=ci("K")), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = v if isinstance(v, (int, float, bool)) else str(v)
    if d:
        data[r] = d

json.dump(data, open(OUT + "rg_rego.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("registration rows", len(data), file=sys.stderr)
for r in sorted(data):
    print(r, data[r], file=sys.stderr)
wb.close()
