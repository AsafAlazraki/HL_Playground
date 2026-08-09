import openpyxl, sys, io, json
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Trailer Module"]

# column indices (1-based)
IDX = dict(A=1, B=2, C=3, D=4, E=5, H=8, I=9, J=10, K=11, L=12, M=13, N=14, O=15,
           AN=40, AQ=43, AR=44, AS=45, BS=71, BT=72, BU=73, BV=74, BW=75,
           BY=77, BZ=78, CA=79)
MAXC = 385
rows = []
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=918, max_col=MAXC, values_only=True), 1):
    rows.append((i, r))
wb.close()

def g(r, letter):
    from openpyxl.utils import column_index_from_string as cifs
    j = cifs(letter) - 1
    return r[j] if j < len(r) else None

brand = None
series = None
recs = []
struct = []
for i, r in rows:
    a = g(r, 'A'); c = g(r, 'C'); e = g(r, 'E'); k = g(r, 'K')
    nonempty = sum(1 for v in r if v is not None and str(v).strip() != '')
    if a is not None and str(a).strip() != '':
        brand = str(a).strip()
        series = str(c).strip() if c else None
        struct.append(('BRAND+SERIES', i, brand, series, nonempty))
        continue
    if (c is not None and str(c).strip() != '') and (e is None or str(e).strip() == '') and nonempty < 8:
        series = str(c).strip()
        struct.append(('SERIES', i, brand, series, nonempty))
        continue
    if c is not None and str(c).strip() != '' and e is not None:
        recs.append(dict(row=i, brand=brand, series=series, name=str(c).strip(), code=str(e).strip(),
                         nonempty=nonempty))
    elif nonempty > 0:
        struct.append(('OTHER', i, brand, series, nonempty))

print("STRUCT ROWS:", len(struct), " DATA ROWS:", len(recs))
print("\n--- BRAND / SERIES HEADER ROWS ---")
for t, i, b, s, n in struct:
    print(f"  r{i:<4} {t:<12} brand={b} | series={s} | cells={n}")

print("\n--- BRAND -> SERIES -> COUNT ---")
from collections import Counter, OrderedDict
agg = OrderedDict()
for d in recs:
    agg.setdefault(d['brand'], OrderedDict()).setdefault(d['series'], []).append(d)
for b, sm in agg.items():
    tot = sum(len(v) for v in sm.values())
    print(f"BRAND: {b}  (series={len(sm)}, trailers={tot})")
    for s, v in sm.items():
        print(f"    - {s}: {len(v)}  e.g. {v[0]['name']} / {v[0]['code']}")

json.dump(recs, open('t05_records.json', 'w', encoding='utf-8'), indent=0)
json.dump([list(x) for x in struct], open('t05_struct.json', 'w', encoding='utf-8'), indent=0)
