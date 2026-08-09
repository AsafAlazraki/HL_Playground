import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
WANT = [1,2,3,143,200,226,233,248,262,278,949,951,955]
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Boat Module"]
cap={}
for i,row in enumerate(ws.iter_rows(min_row=1, max_row=960, values_only=True), start=1):
    if i in WANT:
        cap[i]=[(j,v) for j,v in enumerate(row,1) if v is not None and str(v).strip()!=""]
wb.close()
json.dump({str(k):[[j,str(v)] for j,v in v2] for k,v2 in cap.items()}, open(SP+"bm_headers.json","w"), indent=1)

for r in WANT:
    v=cap.get(r,[])
    print(f"\n===== ROW {r}  ({len(v)} non-empty) =====")
    # compress: skip runs of 'STANDARD FACTORY INCLUSIONS - NN' style
    prev=None
    for j,val in v:
        s=str(val)
        print(f"{gl(j)}{r}\t{s[:120]}")
