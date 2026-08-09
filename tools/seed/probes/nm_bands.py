from openpyxl import load_workbook
SRC=r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
wb=load_workbook(SRC, read_only=True, data_only=True)
ws=wb['Parts Maintenance']
targets={6,17,58,176,226,245,379,380,619,650,736,1097,2472,2542,2847,3185,3201,3217}
last=None; out={}
rn=0
for row in ws.iter_rows(min_row=1,max_row=3722,max_col=5,values_only=True):
    rn+=1
    c=row[2]; e=row[4]
    if isinstance(c,str) and c.strip() and (e in (None,'')):
        last=(rn,c)
    if rn in targets: out[rn]=(c,last)
for r in sorted(out): print(r,'->',out[r][1],'  |',out[r][0])
ws2=wb['Dealer Fit Module']
targets2={32,77,78,96,683,714,1598,1599}
last=None; out2={}; rn=0
for row in ws2.iter_rows(min_row=1,max_row=2680,max_col=4,values_only=True):
    rn+=1
    c=row[2]; d=row[3]
    if isinstance(c,str) and c.strip() and (d in (None,'')) and rn>=12:
        last=(rn,c)
    if rn in targets2: out2[rn]=(c,last)
print('--- DFM')
for r in sorted(out2): print(r,'->',out2[r][1],'  |',out2[r][0])
wb.close()
