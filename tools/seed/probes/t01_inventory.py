import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS:", len(wb.sheetnames))
for n in wb.sheetnames:
    ws = wb[n]
    print(f"  [{n}] rows={ws.max_row} cols={ws.max_column} state={ws.sheet_state}")
wb.close()
