import openpyxl, sys, json, re, collections, time
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"

norm = lambda f: re.sub(r'(?<=[A-Z$])(\d+)', '{r}', f)

# ---------- PASS 1: formulas ----------
t0=time.time()
wb=openpyxl.load_workbook(P, read_only=True, data_only=False)
ws=wb["Boat Module"]
fpat=collections.defaultdict(collections.Counter)   # col -> Counter(pattern)
ftot=0
for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
    for j,v in enumerate(row,1):
        if isinstance(v,str) and v.startswith("="):
            ftot+=1
            fpat[j][norm(v)[:300]] += 1
wb.close()
print("formula pass secs", round(time.time()-t0,1), "total formula cells", ftot, "cols with formulas", len(fpat))
json.dump({gl(j):dict(c.most_common(12)) for j,c in fpat.items()}, open(SP+"bm_formula_patterns.json","w"), indent=1)

# ---------- PASS 2: values / enums / 560 hits ----------
t0=time.time()
wb=openpyxl.load_workbook(P, read_only=True, data_only=True)
ws=wb["Boat Module"]
fill=collections.Counter()
vals=collections.defaultdict(set)
OVER=set()
hits560=[]
pat=re.compile(r'(SP\s?560|Sport\s?560|S560|SP560)', re.I)
for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
    for j,v in enumerate(row,1):
        if v is None: continue
        s=str(v)
        if s.strip()=="": continue
        fill[j]+=1
        if j not in OVER:
            vals[j].add(s[:80])
            if len(vals[j])>150:
                OVER.add(j); vals[j]=set()
        if pat.search(s):
            hits560.append((i,j,s[:200]))
wb.close()
print("value pass secs", round(time.time()-t0,1))
json.dump({"fill":{gl(j):n for j,n in fill.items()},
           "enums":{gl(j):sorted(v) for j,v in vals.items() if j not in OVER},
           "highcard":[gl(j) for j in sorted(OVER)]},
          open(SP+"bm_enums.json","w"), indent=1)
json.dump([[i,gl(j),s] for i,j,s in hits560], open(SP+"bm_560_hits.json","w"), indent=1)
print("560 hits:", len(hits560))
rows560 = sorted(set(i for i,_,_ in hits560))
print("rows containing 560 refs:", len(rows560))
cols560 = collections.Counter(gl(j) for _,j,_ in hits560)
print("cols:", cols560.most_common(40))
