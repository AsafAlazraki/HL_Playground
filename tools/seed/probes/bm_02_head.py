import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Boat Module"]

# find real extent of data: scan first 25 rows, record non-empty cols
maxcol_seen = 0
rows = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25), start=1):
    vals = {}
    for c in row:
        if c.value is not None and str(c.value).strip() != "":
            vals[c.column] = c.value
            if c.column > maxcol_seen: maxcol_seen = c.column
    rows.append((i, vals))
print("max non-empty col in first 25 rows:", maxcol_seen, gl(maxcol_seen))
for i, vals in rows:
    if not vals:
        print(f"R{i}: <empty>")
        continue
    s = " | ".join(f"{gl(k)}{i}={str(v)[:60]}" for k,v in sorted(vals.items())[:60])
    print(f"R{i} (n={len(vals)}): {s}")
wb.close()
