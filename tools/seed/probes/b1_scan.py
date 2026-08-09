import openpyxl, json, sys
from openpyxl.utils import get_column_letter

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Boat Module"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

out = []
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=1010, min_col=1, max_col=22), start=1):
    vals = []
    for c in row:
        v = c.value
        vals.append("" if v is None else str(v)[:60])
    if any(v != "" for v in vals):
        out.append([r] + vals)

with open(r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/b1_rows.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False)
print("ROWS WRITTEN", len(out), file=sys.stderr)
wb.close()
