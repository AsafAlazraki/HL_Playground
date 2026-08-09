import openpyxl, sys, re, collections
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

wbf = openpyxl.load_workbook(SAL, read_only=True, data_only=False)
wsf = wbf['Motor Library']
lit = collections.Counter(); anchor = collections.Counter()
anchor_targets = collections.Counter(); ext = collections.Counter(); nf = 0
LITRE = re.compile(r'^=[0-9.]+([-+*/][0-9.]+)+$')
ANCH = re.compile(r'^=([A-Z]{1,3})\$(\d+)$')
for row in wsf.iter_rows(min_row=1, max_row=660, max_col=651):
    for c in row:
        v = c.value
        if not (isinstance(v, str) and v.startswith('=')):
            continue
        nf += 1
        if LITRE.match(v):
            lit[v] += 1
        m = ANCH.match(v)
        if m:
            anchor[c.column_letter] += 1
            anchor_targets[int(m.group(2))] += 1
        for e in re.findall(r'\[(\d)\]', v):
            ext[e] += 1
wbf.close()
print('total formulas rows1-660:', nf)
print('LITERAL-ARITHMETIC formulas (hard-typed numbers):', sum(lit.values()), 'distinct', len(lit))
for k, n in lit.most_common(16):
    print('   ', n, k)
print('\nABSOLUTE-ROW ANCHOR refs (=X$row) total:', sum(anchor.values()))
print('   target rows (top 22):', anchor_targets.most_common(22))
print('   columns (top 15):', anchor.most_common(15))
print('\nEXTERNAL WORKBOOK ref counts by [index]:', ext.most_common())

wbv = openpyxl.load_workbook(SAL, read_only=True, data_only=True)
wsv = wbv['Motor Library']
err = collections.Counter(); errcol = collections.Counter(); errrow = collections.Counter()
zeros = collections.Counter()
for row in wsv.iter_rows(min_row=1, max_row=660, max_col=651):
    for c in row:
        v = c.value
        if isinstance(v, str) and v.startswith('#'):
            err[v] += 1; errcol[c.column_letter] += 1; errrow[c.row] += 1
        if isinstance(v, str) and v.strip() == '.':
            zeros[c.column_letter] += 1
wbv.close()
print('\nERROR VALUES total:', sum(err.values()), err.most_common())
print('   top error columns:', errcol.most_common(18))
print('   top error rows:', errrow.most_common(12))
print("\n'.' placeholder cells total:", sum(zeros.values()), 'in', len(zeros), 'cols; top:', zeros.most_common(10))
