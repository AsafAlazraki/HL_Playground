import openpyxl, collections, re
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a): lines.append(' '.join(str(x) for x in a))

ERR = ('#N/A','#REF!','#VALUE!','#DIV/0!','#NAME?','#NULL!','#NUM!')
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
p('=== ERROR-VALUE CENSUS (computed values) ===')
for name in wb.sheetnames:
    ws = wb[name]
    cnt = collections.Counter()
    percol = collections.Counter()
    for row in ws.iter_rows(values_only=False if False else True):
        for ci, v in enumerate(row, 1):
            if isinstance(v, str) and v in ERR:
                cnt[v] += 1
                percol[ci] += 1
    if cnt:
        top = ', '.join(f'{get_column_letter(c)}:{n}' for c, n in percol.most_common(8))
        p(f'  {name:28} total={sum(cnt.values()):6} {dict(cnt)}  topcols[{top}]')
wb.close()

p('')
p('=== FILL-COLOUR CENSUS on Parts Maintenance (manual-override markers) ===')
wb2 = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
ws = wb2['Parts Maintenance']
fills = collections.defaultdict(collections.Counter)
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=28):
    for cell in row:
        f = cell.fill
        if f is not None and f.patternType == 'solid':
            try:
                rgb = f.fgColor.rgb
            except Exception:
                rgb = None
            if not isinstance(rgb, str):
                rgb = f'theme{getattr(f.fgColor,"theme",None)}'
            fills[rgb][cell.column] += 1
for rgb, cols in sorted(fills.items(), key=lambda x: -sum(x[1].values()))[:10]:
    tot = sum(cols.values())
    top = ', '.join(f'{get_column_letter(c)}:{n}' for c, n in cols.most_common(8))
    p(f'  {rgb:14} total={tot:6}  [{top}]')

# hardcoded (non-formula) values in normally-computed columns = manual overrides
p('')
p('=== MANUAL OVERRIDES: hardcoded values in computed columns (Parts Maintenance) ===')
COMP = {6:'F SupDesc',7:'G P&A',8:'H MU',9:'I CTD',12:'L Sell',14:'N Install Type',27:'AA Op Code'}
for c, nm in COMP.items():
    hard = 0; form = 0
    for r in range(3, ws.max_row+1):
        v = ws.cell(row=r, column=c).value
        if v is None or (isinstance(v,str) and v.strip()==''): continue
        if isinstance(v, str) and v.startswith('='): form += 1
        else: hard += 1
    p(f'  {nm:16} formula={form:5} hardcoded={hard:5}')
wb2.close()
open(OUT+'p8_pain.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
