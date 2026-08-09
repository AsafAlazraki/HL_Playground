import openpyxl, sys, io, json
from collections import Counter, OrderedDict
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cifs
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"

# --- registration checklist region KE..NU : which rows carry data ---
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Trailer Module"]
lo, hi = cifs('KD'), cifs('NU')
hits = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=918, min_col=lo, max_col=hi, values_only=True), 1):
    cells = [(gcl(lo+j), v) for j, v in enumerate(row) if v is not None and str(v).strip() != '']
    if cells:
        hits.append((i, cells))
print("=== REGION KD..NU  rows with content:", len(hits))
for i, cells in hits[:14]:
    print(f" r{i}: n={len(cells)}")
    print("    " + " | ".join(f"{a}={str(b)[:38]}" for a, b in cells[:40]))
# columns GT..IV
lo2, hi2 = cifs('GT'), cifs('IV')
h2 = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=918, min_col=lo2, max_col=hi2, values_only=True), 1):
    cells = [(gcl(lo2+j), v) for j, v in enumerate(row) if v is not None and str(v).strip() != '']
    if cells: h2.append((i, cells))
print("\n=== REGION GT..IV rows with content:", len(h2))
for i, cells in h2[:6]:
    print(f" r{i}: " + " | ".join(f"{a}={str(b)[:40]}" for a, b in cells[:25]))
wb.close()

# --- fills / fonts on header row + a data row ---
wb = openpyxl.load_workbook(P, read_only=False, data_only=True)
ws = wb["Trailer Module"]
print("\n=== FILL COLOURS: header row 1 (C..CA) ===")
runs = []
for j in range(cifs('C'), cifs('CA')+1):
    c = ws.cell(row=1, column=j)
    fg = c.fill.start_color.rgb if c.fill and c.fill.start_color else None
    fg = fg if isinstance(fg, str) else None
    b = c.font.bold if c.font else None
    runs.append((gcl(j), str(c.value)[:26] if c.value else '', fg, b))
cur = None
for L, v, fg, b in runs:
    if fg != cur:
        print(f"  --- fill {fg} starts at {L} ({v})")
        cur = fg
print("\n=== FILL COLOURS on a data row (r5) ===")
cur = None
for j in range(cifs('C'), cifs('CA')+1):
    c = ws.cell(row=5, column=j)
    fg = c.fill.start_color.rgb if c.fill and c.fill.start_color else None
    fg = fg if isinstance(fg, str) else None
    if fg != cur:
        print(f"  --- fill {fg} starts at {gcl(j)}  hdr={str(ws.cell(row=1,column=j).value)[:30]}")
        cur = fg
print("\n=== BRAND BANNER ROW FILLS ===")
for r in (186, 232, 281, 455, 626, 656, 4, 42, 187):
    c = ws.cell(row=r, column=3)
    a = ws.cell(row=r, column=1)
    fg = c.fill.start_color.rgb if c.fill and c.fill.start_color else None
    print(f"  r{r}: A={str(a.value)[:34]:<34} C={str(c.value)[:40]:<40} fill={fg} bold={c.font.bold} size={c.font.size}")
wb.close()
