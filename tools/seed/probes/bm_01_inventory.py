import openpyxl, sys, io, json
sys.stdout.reconfigure(encoding='utf-8')

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS:", len(wb.sheetnames))
out=[]
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  [{name}] state={ws.sheet_state} dims={ws.calculate_dimension()} max_row={ws.max_row} max_col={ws.max_column}")
    out.append({"name":name,"state":ws.sheet_state,"max_row":ws.max_row,"max_col":ws.max_column})
wb.close()
json.dump(out, open(r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/bm_sheets.json","w"), indent=1)
