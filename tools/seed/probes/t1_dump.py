import openpyxl, json, sys
from openpyxl.utils import get_column_letter as gl

P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
OUT = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Trailer Module"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
style = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=720, min_col=1, max_col=80), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = v if isinstance(v, (int, float, bool)) else str(v)
        if i == 3 and v is not None and str(v).strip() != "":
            try:
                style[r] = {"sz": c.font.sz, "b": c.font.b}
            except Exception as e:
                style[r] = {"err": str(e)}
    if d:
        data[r] = d

json.dump(data, open(OUT + "t1_data.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
json.dump(style, open(OUT + "t1_style.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("rows", len(data), file=sys.stderr)

ws2 = wb["Dropdowns"]
dd = {}
for r, row in enumerate(ws2.iter_rows(min_row=1, max_row=60, min_col=1, max_col=30), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = str(v)
    if d:
        dd[r] = d
json.dump(dd, open(OUT + "t1_dropdowns.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("dropdown rows", len(dd), file=sys.stderr)
wb.close()
