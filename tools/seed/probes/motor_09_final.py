import openpyxl, sys, collections, json, re
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

wbv = openpyxl.load_workbook(SAL, read_only=True, data_only=True)
ws = wbv['Motor Library']
V = [list(r) + [None]*(651-len(r)) for r in ws.iter_rows(min_row=1, max_row=660, max_col=651, values_only=True)]
wbv.close()
wbf = openpyxl.load_workbook(SAL, read_only=True, data_only=False)
wsf = wbf['Motor Library']
F = [list(r) + [None]*(651-len(r)) for r in wsf.iter_rows(min_row=1, max_row=660, max_col=651, values_only=True)]
wbf.close()
hdr = V[3]

def nonblank(v): return v is not None and str(v).strip() not in ('', '.')

# --- classify rows
banner, datar = [], []
for i in range(5, 661):
    r = V[i-1]
    n = sum(1 for c in range(651) if nonblank(r[c]))
    if n == 0: continue
    if n <= 3 and nonblank(r[2]): banner.append((i, str(r[2])))
    else: datar.append(i)
print("TRUE BANNERS:", len(banner), " TRUE DATA ROWS:", len(datar), "range", datar[0], "-", datar[-1])
for i, s in banner: print(f"   r{i}: {s}")

# --- section map: assign each data row to the last banner above it
sec = collections.OrderedDict()
bl = None; bi = 0
for i in datar:
    while bi < len(banner) and banner[bi][0] < i:
        bl = banner[bi][1]; bi += 1
    sec.setdefault(bl, []).append(i)
print("\n=== SECTION -> row range, count ===")
for k, v in sec.items():
    print(f"   {str(k)[:70]:<70} r{v[0]}-r{v[-1]}  n={len(v)}")

# --- manual vs formula per key column
print("\n=== MANUAL vs FORMULA (data rows) ===")
for L in ['C','D','E','F','G','H','I','J','K','L','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AC','AJ','AV','BB','BC','BD','BE','BF','BJ','BK','BL','BQ','BR','BS','BW','BX','BY','CI','DA','DB','EY','EZ','FA','FB','FC','FF','GT','GU','LC','LE','LI','LK','LO','LT','LV','TZ','WS']:
    c = ci(L)-1
    val = sum(1 for i in datar if nonblank(V[i-1][c]))
    fml = sum(1 for i in datar if isinstance(F[i-1][c], str) and str(F[i-1][c]).startswith('='))
    print(f"   {L:<3} {str(hdr[c])[:28]:<28} filled={val:>4}  formula={fml:>4}  MANUAL={val-fml:>4}")

# --- rigging / prop option fill depth
print("\n=== OPTION SLOT FILL DEPTH ===")
for tag, a, b in [('Rigging Option', 'DA', 'EX'), ("Additional FO's", 'FT', 'GR'), ('Prop Option', 'GT', 'KO')]:
    tot = 0; per = []
    for i in datar:
        n = sum(1 for c in range(ci(a)-1, ci(b)) if nonblank(V[i-1][c]))
        per.append(n); tot += n
    per_s = sorted(per)
    print(f"   {tag} [{a}:{b}] slots={ci(b)-ci(a)+1} used_cells={tot} max_per_row={max(per)} median={per_s[len(per_s)//2]} rows_with_0={per.count(0)}")

# --- shaft length x rotation x model-code prefix taxonomy
print("\n=== SHAFT/ROTATION/CODE taxonomy ===")
sh = collections.Counter(); lpref = collections.Counter(); comb = collections.Counter()
for i in datar:
    f = V[i-1][ci('F')-1]; d = V[i-1][ci('D')-1]
    if nonblank(f): sh[str(f)] += 1
    if nonblank(d):
        s = str(d)
        lpref['L-prefix (counter-rotating)' if s.startswith('L') and not s.startswith('LWK') else 'standard'] += 1
        comb[(str(f), s.startswith('L') and not s.startswith('LWK'))] += 1
print("   Shaft Length enum:", sh.most_common())
print("   Model code L-prefix:", lpref.most_common())

# --- suffix analysis on model code (last chars = generation letter)
suf = collections.Counter()
for i in datar:
    d = V[i-1][ci('D')-1]
    if nonblank(d):
        m = re.match(r'^(L?)(V?X?F|T)([\d.]+)([A-Z]{1,3})(\d?)', str(d))
        if m: suf[(m.group(1) or '-', m.group(2), m.group(4), m.group(5) or '-')] += 1
print("   code pattern (Lprefix, family, letters, trailing digit) top:", suf.most_common(24))

# --- Haines Signature powerplant row 345 vs base F115XB row 90
print("\n=== POWERPLANT ROW r345 'SIG 525F w Yamaha - F115XB' key cells ===")
for L in ['C','D','E','F','G','H','J','K','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AC','AV','AX','AY','AZ','BB','BC','BD','BE','BF','BL','CI','DA','GT','TZ']:
    c = ci(L)-1
    print(f"   {L}345 [{str(hdr[c])[:24]}] val={str(V[344][c])[:70]!r}  f={str(F[344][c])[:90] if isinstance(F[344][c],str) and str(F[344][c]).startswith('=') else ''}")
