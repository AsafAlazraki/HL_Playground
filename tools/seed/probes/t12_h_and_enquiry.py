import json, sys, io, re, openpyxl
from collections import Counter
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
recs = json.load(open('t07_records.json'))
c = Counter(str(r.get('H')) for r in recs if r.get('H') not in (None,'None'))
ks = sorted(c, key=str)
print("=== H (Boat Size) ALL DISTINCT (n=%d) ===" % len(ks))
print(" | ".join(f"{k}({c[k]})" for k in ks))
def kind(s):
    s = s.strip()
    if re.fullmatch(r'\d+(\.\d+)?', s): return 'plain number (metres or model no.)'
    if re.fullmatch(r'\d+(\.\d+)?\s*[Mm]', s): return 'number + M suffix'
    if re.search(r'\d\s*[-/]\s*\d', s): return 'RANGE'
    return 'TEXT / boat model name'
kc = Counter(kind(k) for k in ks)
print("\n=== H VALUE SHAPES (distinct-value basis) ===")
for k, v in kc.most_common(): print(f"   {k}: {v}")

# Trailer Spec Enquiry full layout
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
for do in (True, False):
    wb = openpyxl.load_workbook(P, read_only=True, data_only=do)
    ws = wb["Trailer Spec Enquiry"]
    print("\n" + "="*90)
    print("TRAILER SPEC ENQUIRY  (data_only=%s)" % do)
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=91, max_col=40, values_only=True), 1):
        cells = [(gcl(j+1), v) for j, v in enumerate(r) if v is not None and str(v).strip() != '']
        if cells:
            print(f" r{i:<3}: " + " | ".join(f"{a}={str(b)[:60]}" for a, b in cells))
    wb.close()
