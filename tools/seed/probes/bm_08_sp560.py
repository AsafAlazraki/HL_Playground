import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
H = json.load(open(SP+"bm_headers.json"))
hdr = {int(j):v for j,v in H["278"]}

ROWS = list(range(276,300)) + list(range(825,850)) + [1436,1437] + list(range(1531,1552))
res={}
for mode in (True, False):
    wb = openpyxl.load_workbook(P, read_only=True, data_only=mode)
    ws = wb["Boat Module"]
    d={}
    for i,row in enumerate(ws.iter_rows(min_row=1, max_row=1560, values_only=True), start=1):
        if i in ROWS:
            d[i] = {j:v for j,v in enumerate(row,1) if v is not None and str(v).strip()!=""}
    wb.close()
    res["values" if mode else "formulas"]=d

# save full json
json.dump({k:{str(r):{gl(c):str(v) for c,v in cols.items()} for r,cols in dd.items()} for k,dd in res.items()},
          open(SP+"bm_sp560_raw.json","w"), indent=1)

V=res["values"]; F=res["formulas"]
def dump(r, cmin=1, cmax=700, show_formula=True):
    print(f"\n########## ROW {r}  (C={V.get(r,{}).get(3)}) ##########")
    cols = V.get(r,{})
    for c in sorted(cols):
        if not (cmin<=c<=cmax): continue
        h = hdr.get(c,"")
        val = cols[c]
        f = F.get(r,{}).get(c)
        fs = ""
        if show_formula and isinstance(f,str) and f.startswith("="):
            fs = f"   <<FORMULA {f[:160]}>>"
        print(f"  {gl(c)}{r} [{h}] = {str(val)[:200]}{fs}")

for r in [1533, 1534, 1535, 1548]:
    dump(r)
