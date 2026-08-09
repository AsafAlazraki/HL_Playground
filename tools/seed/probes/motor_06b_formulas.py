import re, sys, collections, json, openpyxl
from openpyxl.utils import column_index_from_string as cidx
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

hdrmap = {}
for line in open('motor_header_grid.txt', encoding='utf8'):
    p = line.rstrip('\n').split('\t')
    if len(p) > 4:
        hdrmap[p[0]] = p[4]

wb = openpyxl.load_workbook(SAL, read_only=True, data_only=False)
allpat = {}
for ws in wb.worksheets:
    bycol = collections.defaultdict(collections.Counter)
    n = 0
    maxr = min(ws.max_row, 700) if ws.title == 'Motor Library' else ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=maxr):
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith('='):
                rn = str(c.row)
                norm = re.sub(r'(?<![0-9])' + re.escape(rn) + r'(?![0-9])', '{r}', v)
                bycol[c.column_letter][norm] += 1
                n += 1
    allpat[ws.title] = {c: dict(v.most_common(8)) for c, v in bycol.items()}
    print(f"\n================ {ws.title}: {n} formula cells across {len(bycol)} cols")
    for c in sorted(bycol, key=cidx):
        h = hdrmap.get(c, '') if ws.title == 'Motor Library' else ''
        print(f"  [{c}] {h[:32]:<32} n={sum(bycol[c].values()):>5} uniq={len(bycol[c]):>4}")
        for f, k in bycol[c].most_common(3):
            print(f"        ({k:>4}) {f[:250]}")
wb.close()
json.dump(allpat, open('motor_formula_patterns.json', 'w'), indent=1)
