import zipfile, re, sys, collections, openpyxl
from openpyxl.utils import get_column_letter as gl
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"
z = zipfile.ZipFile(SAL)
ss = z.read('xl/sharedStrings.xml').decode('utf8')
strings = re.findall(r'<si>(.*?)</si>', ss, re.S)
strings = [re.sub(r'<[^>]+>', '', s) for s in strings]
print("sharedStrings:", len(strings))
KW = ['max hp', 'maximum hp', 'rated hp', 'transom', 'fitment', 'compatib', 'envelope',
      'min hp', 'suits', 'boat model', 'hull', 'recommend', 'approved']
for k in KW:
    hits = [s for s in strings if k in s.lower()]
    print(f"\n  KEYWORD '{k}': {len(hits)}")
    for h in hits[:8]:
        print("     ", h.replace('\n', ' ')[:150])
z.close()

# column A/B content and colour scan
wb = openpyxl.load_workbook(SAL, data_only=True)
ws = wb['Motor Library']
print("\n=== A/B columns rows 1..12 ===")
for r in range(1, 13):
    print(f"  r{r}: A={ws.cell(r,1).value!r} B={ws.cell(r,2).value!r}")

print("\n=== FILL COLOURS on header row 4 (col -> rgb) ===")
fills = collections.Counter()
colfill = {}
for c in range(1, 652):
    cell = ws.cell(4, c)
    f = cell.fill
    rgb = f.start_color.rgb if f and f.fill_type == 'solid' else None
    colfill[gl(c)] = rgb
    fills[str(rgb)] += 1
print("  header-row fill histogram:", fills.most_common(12))
runs = []
prev = None
for c in range(1, 652):
    L = gl(c); v = colfill[L]
    if v != prev:
        runs.append([L, L, v]); prev = v
    else:
        runs[-1][1] = L
print("  header fill runs (col-range -> rgb):")
for a, b, v in runs:
    print(f"     {a}:{b}  {v}")

print("\n=== FILL COLOURS on a data row (r90 Yamaha F115XB) ===")
runs = []; prev = None
for c in range(1, 652):
    cell = ws.cell(90, c)
    f = cell.fill
    rgb = f.start_color.rgb if f and f.fill_type == 'solid' else None
    L = gl(c)
    if rgb != prev:
        runs.append([L, L, rgb]); prev = rgb
    else:
        runs[-1][1] = L
for a, b, v in runs[:70]:
    print(f"     {a}:{b}  {v}")
wb.close()
