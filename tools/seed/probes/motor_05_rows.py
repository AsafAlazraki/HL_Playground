import openpyxl, sys, json
from openpyxl.utils import get_column_letter, column_index_from_string as ci
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"
wb = openpyxl.load_workbook(SAL, read_only=True, data_only=True)
ws = wb['Motor Library']
data = [list(r) + [None]*(651-len(r)) for r in ws.iter_rows(min_row=1, max_row=660, max_col=651, values_only=True)]
wb.close()
hdr = data[3]

def dump(rownum, maxcols=651, label=""):
    r = data[rownum-1]
    print(f"\n########## ROW {rownum} {label}")
    for c in range(maxcols):
        v = r[c]
        if v is None or (isinstance(v, str) and v.strip() in ('', '.')):
            continue
        h = hdr[c]
        L = get_column_letter(c+1)
        s = str(v).replace('\n', ' | ')
        print(f"   {L}{rownum} [{h}] = {s[:150]}")

# 1) A representative simple Yamaha row + the F115 rows
print("=== SEARCH for F115 ===")
for i in range(5, 660):
    d = data[i-1][ci('D')-1]
    if d and 'F115' in str(d):
        print(f"  r{i}: C={data[i-1][ci('C')-1]!r} D={d!r} E={data[i-1][ci('E')-1]!r} F={data[i-1][ci('F')-1]!r}")

print("\n=== 'Powerplants' block row shapes: which columns are filled ===")
for rownum in (342, 343, 345, 348, 349, 393, 394, 431, 432, 449, 450, 515, 611, 612, 623, 624, 629, 630):
    r = data[rownum-1]
    filled = [get_column_letter(c+1) for c in range(651) if r[c] is not None and str(r[c]).strip() not in ('', '.')]
    print(f"  r{rownum}: nfilled={len(filled)} C={str(r[ci('C')-1])[:60]!r} cols={filled[:40]}")
