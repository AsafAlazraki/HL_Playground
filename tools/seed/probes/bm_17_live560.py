import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
H=json.load(open(SP+"bm_headers.json")); hdr={gl(int(j)):v for j,v in H["278"]}; hdr1={gl(int(j)):v for j,v in H["1"]}
HH=lambda c: hdr.get(c) or hdr1.get(c,'')
ROWS=set(range(825,850))|{1773,281,282,283,948,947}
res={}
for mode in (True,False):
    wb=openpyxl.load_workbook(P, read_only=True, data_only=mode); ws=wb["Boat Module"]
    d={}
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=1780,values_only=True),start=1):
        if i in ROWS: d[i]={gl(j):v for j,v in enumerate(row,1) if v is not None and str(v).strip()!=""}
    wb.close(); res["v" if mode else "f"]=d
json.dump({k:{str(r):{c:str(v) for c,v in cc.items()} for r,cc in dd.items()} for k,dd in res.items()},
          open(SP+"bm_live560_raw.json","w"), indent=1)
V=res["v"];F=res["f"]
SP560=[829,830,831,832,833,834,835,837,838,839,840,841,842,843,844]
print("=== LIVE Highfield 2026 SP560 variants (rows 829-844) ===")
KEYS=["C","D","F","G","H","I","K","L","M","O","P","Q","S","T","X","BY","BZ","CA","CB","CC","CD","CE","CF","CG",
      "II","IJ","IK","IM","IN","IO","IP","IQ","IR","IS","IT","IU","IV","IW","IX","IY",
      "JF","JG","JH","JI","JJ","JK","JM","JN","JO","JP","JQ","JR","JT","JU","JV","JW",
      "KE","KF","KI","KJ","KK","KM","KN","KP","KQ","KR","KS","KT",
      "KV","KW","KX","KY","KZ","LA","LB","LC","LD","LF","LG","LL","LR","LX",
      "NZ","OA","OL","OM","ON","OO","OP",
      "QD","QE","QF","QG","QH","QK","QL","QM","QN","QO",
      "QR","QS","QT","QU","QV","QW","QX","QY","QZ","RA","RB","RC","RN","RO","RP","RQ",
      "SV","SW","SX","SY","SZ","TA","TB","TC","TD","TE","TF","TH","TI","TJ","TK","TO","TP","TQ","TR","TT","TU","TV","TW","UA","UB","UC","UD","UF","UG","UH","UI","UJ"]
for k in KEYS:
    vals=[(r, V.get(r,{}).get(k)) for r in SP560]
    if all(v is None for _,v in vals): continue
    uniq={str(v) for _,v in vals}
    f=F.get(829,{}).get(k)
    fs=f"   FORMULA(R829)={f[:120]}" if isinstance(f,str) and f.startswith("=") else ""
    if len(uniq)==1:
        print(f"\n{k} [{HH(k)}] UNIFORM = {str(vals[0][1])[:300]}{fs}")
    else:
        print(f"\n{k} [{HH(k)}] VARIES:{fs}")
        for r,v in vals: print(f"    R{r} ({V.get(r,{}).get('D')}): {str(v)[:160]}")
print("\n=== row 1773 (purple) ===", {c:str(v)[:60] for c,v in list(V.get(1773,{}).items())[:12]})
print("\n=== rows 281-283 (first 2026-range rows) ===")
for r in [281,282,283]:
    d=V.get(r,{}); print(f"  R{r}: C={d.get('C')} D={d.get('D')} QR={d.get('QR')} IM={d.get('IM')}")
print("\n=== rows 947,948 (last of 2026 range) ===")
for r in [947,948]:
    d=V.get(r,{}); print(f"  R{r}: C={d.get('C')} D={d.get('D')} QR={d.get('QR')}")
