import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci
SRC=r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
COLS="C D E F G H I J K L M N O P Q R S T U V W X Y AA AB".split()
IDX={ci(c):c for c in COLS}; MAXC=max(IDX)
NEEDLES=['SP560','SP520/560','Tube Covers to suit PVC Boat - 5.6','Tube Covers to suit Hypalon Boat - 5.6',
 'GX750B','Apollo RA670','EchoMap Ultra 2 125sv','MFM50','Battery Terminals (2 pairs)','Battery Box - Small',
 'Pre Delivery Allowance (2 hrs)','Sand Anchor Kit - 13lb','Inflatable PFD - Adult 150N (Manual)',
 'Rego Decals (Std) t/s PVC Tubes','Rego Decals (Std) t/s Hypalon Tubes','Safety Gear - Partially Smooth Waters']
wb=load_workbook(SRC, read_only=True, data_only=True)
ws=wb['Parts Maintenance']
out={}
rn=0
for row in ws.iter_rows(min_row=1,max_row=3722,max_col=MAXC,values_only=True):
    rn+=1
    c=row[2] if len(row)>2 else None
    if not isinstance(c,str): continue
    if not any(n.lower() in c.lower() for n in NEEDLES): continue
    cells={}
    for i,v in enumerate(row,start=1):
        L=IDX.get(i)
        if L and v not in (None,''): cells[L]=v
    out[rn]=cells
wb.close()
json.dump(out,open('nm_parts.json','w',encoding='utf-8'),indent=1,default=str)
for r in sorted(out,key=int):
    c=out[r]
    print(r,'|',c.get('C'),'| sup',c.get('D'),'| code',c.get('E'),'| cost G',c.get('G'),'| CTD I',c.get('I'),'| Sell L',c.get('L'),'| N',c.get('N'),'| TTF O',c.get('O'),'| Y',c.get('Y'),'| AA',c.get('AA'))
