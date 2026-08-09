import openpyxl, sys, collections
from openpyxl.utils import get_column_letter as gl
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"
wb = openpyxl.load_workbook(SAL, read_only=True, data_only=True)

for name in ['Rebate Programs', 'Data Drop']:
    ws = wb[name]
    print(f"\n########## {name}  {ws.max_row}x{ws.max_column}")
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [f"{gl(j+1)}{i}={v}" for j, v in enumerate(r) if v is not None]
        if cells: print("  " + " | ".join(str(c)[:120] for c in cells))

ws = wb['Yamaha_Dealer_Current']
print(f"\n########## Yamaha_Dealer_Current {ws.max_row}x{ws.max_column}")
rows = list(ws.iter_rows(values_only=True))
for i, r in enumerate(rows[:14], 1):
    print(f"  r{i}:", [str(v)[:38] for v in r if v is not None][:14])
print("  ...")
for i, r in enumerate(rows[190:215], 191):
    print(f"  r{i}:", [str(v)[:38] for v in r if v is not None][:14])

ws = wb['Dropdowns']
print(f"\n########## Dropdowns {ws.max_row}x{ws.max_column}")
drows = [list(r) + [None]*(18-len(r)) for r in ws.iter_rows(min_row=1, max_row=2094, max_col=18, values_only=True)]
for c in range(18):
    vals = [r[c] for r in drows if r[c] is not None and str(r[c]).strip() != '']
    if not vals: continue
    cnt = collections.Counter(str(v) for v in vals)
    print(f"\n  COL {gl(c+1)}: n={len(vals)} distinct={len(cnt)}  head={[str(v)[:60] for v in vals[:6]]}")
    if len(cnt) <= 30:
        for v, k in cnt.most_common():
            print(f"        {k:>4}  {v[:100]}")
wb.close()
