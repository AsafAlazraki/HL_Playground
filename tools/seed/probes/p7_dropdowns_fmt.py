import openpyxl, collections
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a): lines.append(' '.join(str(x) for x in a))

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ---- Dropdowns column map ----
ws = wb['Dropdowns']
dd = [r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20, values_only=True)]
p('='*100); p('DROPDOWNS SHEET COLUMN MAP (each column = one picklist)')
for c in range(1, 21):
    vals = [r[c-1] for r in dd if c-1 < len(r)]
    nz = [v for v in vals if v is not None and str(v).strip() != '']
    if not nz: continue
    first = str(nz[0])[:70]
    p(f'  COL {get_column_letter(c)}: entries={len(nz)} distinct={len(set(str(v) for v in nz))} first={first!r}')
    p('     head: ' + ' | '.join(str(v)[:38] for v in nz[:6]))
    p('     tail: ' + ' | '.join(str(v)[:38] for v in nz[-4:]))

# ---- TUBE COVERS fitment band ----
ws = wb['Parts Maintenance']
pm = [r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=30, values_only=True)]
p('')
p('='*100); p('PARTS MAINTENANCE R595-660 "TUBE COVERS - To suit Highfield Boats" (model-name fitment in Description)')
for i in range(595, 661):
    r = pm[i-1]
    if r[2] is None: continue
    p(f'  R{i}: C={str(r[2])[:78]} | D={r[3]} | E={r[4]} | L(Sell)={r[11]} | N={str(r[13])[:38]}')

p('')
p('='*100); p('PARTS MAINTENANCE R2759-2800 SUBLET/LABOUR sample')
for i in range(2759, 2801):
    r = pm[i-1]
    if r[2] is None: continue
    p(f'  R{i}: C={str(r[2])[:70]} | D={r[3]} | E={r[4]} | G={r[6]} | L={r[11]} | N={str(r[13])[:40]} | O={r[14]} | Y={r[24]}')

# ---- Dealer Fit tube cover options band ----
ws = wb['Dealer Fit Module']
df = [r for r in ws.iter_rows(min_row=1, max_row=760, max_col=60, values_only=True)]
p('')
p('='*100); p('DEALER FIT R660-723 "TUBE COVER OPTIONS - To suit Highfield Boats"')
for i in range(660, 724):
    r = df[i-1]
    if r[2] is None: continue
    p(f'  R{i}: C={str(r[2])[:74]} | Code={r[3]} | ActSell={r[17]} | slot1={str(r[19])[:44]}')
wb.close()

# ---- formatting signals on Parts Maintenance header + a few rows ----
wb2 = openpyxl.load_workbook(SRC, read_only=False, data_only=True)
ws = wb2['Parts Maintenance']
p('')
p('='*100); p('FORMAT SIGNALS - Parts Maintenance header row 1 + sample data row 15')
for c in range(3, 30):
    h = ws.cell(row=1, column=c)
    d = ws.cell(row=15, column=c)
    def f(cell):
        fill = cell.fill
        fg = getattr(fill.fgColor, 'rgb', None) if fill and fill.patternType else None
        return f'fill={fill.patternType}:{fg} bold={cell.font.bold} color={getattr(cell.font.color,"rgb",None)} fmt={cell.number_format[:18]}'
    p(f'  {get_column_letter(c)} HDR[{str(h.value)[:22]}] {f(h)}')
    p(f'     R15 {f(d)}')
ws = wb2['Dealer Fit Module']
p('')
p('FORMAT SIGNALS - Dealer Fit row 11 header + row 12')
for c in list(range(3, 30)):
    h = ws.cell(row=11, column=c); d = ws.cell(row=12, column=c)
    def f(cell):
        fill = cell.fill
        fg = getattr(fill.fgColor, 'rgb', None) if fill and fill.patternType else None
        return f'fill={fill.patternType}:{fg} bold={cell.font.bold} fmt={cell.number_format[:16]}'
    p(f'  {get_column_letter(c)} HDR[{str(h.value)[:20]}] {f(h)} || R12 {f(d)}')
wb2.close()

open(OUT+'p7_dropdowns_fmt.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
