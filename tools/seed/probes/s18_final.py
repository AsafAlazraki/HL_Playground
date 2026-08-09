import sys; sys.path.insert(0,r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad")
from lib import *
from s8_formulas import formulas, norm
import collections, re, json
BR=["Stabicraft","Surtees","Stacer","Formosa","Jeanneau","Haines Signature","Highfield","Boat Show"]
HDR={"Stabicraft":7,"Surtees":7,"Stacer":7,"Formosa":7,"Jeanneau":7,"Haines Signature":7,"Highfield":7,"Boat Show":10}
# ---- distinct layouts
print("=== BRAND SHEET LAYOUT SIGNATURES ==="); sig={}
for b in BR:
    V,F,mr,mc=grid(b); h=HDR[b]
    s=tuple((get_column_letter(j), (str(V[h][j]) if (j<len(V[h]) and V[h][j] is not None) else None)) for j in range(3,14))
    sig.setdefault(s,[]).append(b)
    print("  %-17s maxcol=%-3d hdr(C..M)=%s"%(b,mc,[x[1] for x in s]))
print("\n  DISTINCT LAYOUTS (C..M header signature): %d"%len(sig))
for s,bs in sig.items(): print("     %s  <- %s"%([x[1] for x in s],bs))

# ---- #REF in dropdowns col B
fs=formulas("Dropdowns")
ref=[(c,r) for (c,r),(f,a) in fs.items() if f and "#REF" in f]
print("\n=== Dropdowns broken #REF! formulas: %d ==="%len(ref))
print("   rows:", sorted(r for c,r in ref)[:45])

# ---- Quote Sheet anchor drift
print("\n=== QUOTE SHEET ANCHOR DRIFT (per-unit cell addresses used) ===")
anch=collections.defaultdict(collections.Counter)
for b in BR:
    for (c,r),(f,a) in formulas(b).items():
        m=re.match(r"'\[(\d+)\](Quote Sheet)'!\$([A-Z]+)\$(\d+)$", f or "")
        if m: anch[c][m.group(3)+m.group(4)]+=1
for c in sorted(anch):
    print("   col %s: %s"%(c, dict(anch[c])))

# ---- merged ranges all sheets
import openpyxl
wb=openpyxl.load_workbook(r"C:/Users/AsafA/Downloads/MASTER PRICE FILE.xlsx", data_only=True)
print("\n=== MERGED RANGES WITH THEIR ANCHOR VALUE ===")
for n in wb.sheetnames:
    ws=wb[n]; rs=sorted(str(x) for x in ws.merged_cells.ranges)
    if not rs: continue
    print("\n  [%s] %d"%(n,len(rs)))
    for r in rs:
        a=r.split(":")[0]
        v=ws[a].value
        print("     %-12s = %s"%(r, repr(v)[:90]))
wb.close()
