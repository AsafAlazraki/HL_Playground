import openpyxl, zipfile, re
from openpyxl.utils import get_column_letter
SRC = r"C:/Users/AsafA/Downloads/MASTER PRICE FILE.xlsx"
wv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
wf = openpyxl.load_workbook(SRC, read_only=True, data_only=False)

def dump(sheet, maxr, maxc, label=""):
    ws=wv[sheet]; wsf=wf[sheet]
    rowsv=list(ws.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True))
    rowsf=list(wsf.iter_rows(min_row=1,max_row=maxr,max_col=maxc,values_only=True))
    print(f"\n########## {sheet} {label} rows1-{maxr} cols1-{maxc} ##########")
    for i,(rv,rf) in enumerate(zip(rowsv,rowsf), start=1):
        cells=[]
        for j,(v,f) in enumerate(zip(rv,rf), start=1):
            if v is None and f is None: continue
            L=get_column_letter(j)
            s=f"{L}{i}="
            if isinstance(f,str) and f.startswith("="):
                s+=f"[F]{f[:180]}"
                if v is not None: s+=f" ->{v!r}"[:80]
            else:
                s+=repr(v)[:160]
            cells.append(s)
        if cells: print(f" r{i}: " + " | ".join(cells))

dump("Home Page",44,16)
wv.close(); wf.close()

# hyperlinks on Home Page
z=zipfile.ZipFile(SRC)
xml=z.read("xl/worksheets/sheet1.xml").decode("utf8","replace")
print("\n=== Home Page hyperlinks ===")
for m in re.finditer(r'<hyperlink [^>]*/>', xml): print("  ", m.group(0))
try:
    r=z.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf8","replace")
    for m in re.finditer(r'<Relationship[^>]*>', r):
        if "hyperlink" in m.group(0): print("  REL:", m.group(0)[:400])
except KeyError: pass
