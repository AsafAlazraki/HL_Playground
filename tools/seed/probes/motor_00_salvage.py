"""Salvage the truncated 'Motor Module (1).xlsx' into a valid xlsx in the SCRATCHPAD.
SOURCE IS READ-ONLY. Nothing in Downloads is touched.
"""
import struct, zlib, zipfile, os, re, sys
sys.stdout.reconfigure(encoding='utf-8')

SRC = r"C:/Users/AsafA/Downloads/Motor Module (1).xlsx"
OUT = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

data = open(SRC, 'rb').read()
entries = {}
off = 0
while True:
    i = data.find(b'PK\x03\x04', off)
    if i < 0:
        break
    try:
        sig, ver, flag, meth, mt, md, crc, csz, usz, fnl, efl = struct.unpack('<IHHHHHIIIHH', data[i:i+30])
        fn = data[i+30:i+30+fnl].decode('utf8', 'replace')
        if not re.match(r'^[\w\[\]\./_ -]+$', fn) or csz == 0:
            off = i+4; continue
        start = i + 30 + fnl + efl
        blob = data[start:start+csz]
        if len(blob) < csz:
            print("TRUNCATED, dropping:", fn, len(blob), "of", csz)
            off = i+4; continue
        raw = zlib.decompress(blob, -15) if meth == 8 else blob
        entries[fn] = raw
    except Exception as e:
        pass
    off = i + 4

print("recovered entries:", len(entries))
for k, v in entries.items():
    print(f"   {k:55s} {len(v):>10,}")

# --- report what the external links say BEFORE we strip them ---
for k in sorted(entries):
    if 'externalLink' in k:
        t = entries[k].decode('utf8', 'replace')
        names = re.findall(r'<sheetName val="([^"]*)"', t)
        print(f"\n[{k}] len={len(t)} sheetNames={names[:20]}")
        print("   head:", t[:400].replace('\n', ' '))

# --- patch workbook.xml / rels / content types to drop missing externalLinks ---
present = {int(m.group(1)) for k in entries if (m := re.match(r'xl/externalLinks/externalLink(\d+)\.xml$', k))}
print("\nexternalLink parts present:", sorted(present))

rels = entries['xl/_rels/workbook.xml.rels'].decode('utf8')
print("\nworkbook.xml.rels:\n", rels[:4000])
wbx = entries['xl/workbook.xml'].decode('utf8')
print("\nworkbook.xml:\n", wbx[:4000])

# find rIds of external links that are missing
missing_rids = []
for m in re.finditer(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="externalLinks/externalLink(\d+)\.xml"[^>]*/>', rels):
    if int(m.group(2)) not in present:
        missing_rids.append(m.group(1))
print("missing rIds:", missing_rids)

for rid in missing_rids:
    rels = re.sub(r'<Relationship[^>]*Id="%s".*?/>' % re.escape(rid), '', rels)
    wbx = re.sub(r'<externalReference[^>]*r:id="%s"[^>]*/>' % re.escape(rid), '', wbx)
wbx = re.sub(r'<externalReferences\s*/>', '', wbx)
wbx = re.sub(r'<externalReferences>\s*</externalReferences>', '', wbx)

ct = entries['[Content_Types].xml'].decode('utf8')
for n in re.findall(r'/xl/externalLinks/externalLink(\d+)\.xml', ct):
    if int(n) not in present:
        ct = ct.replace('<Override PartName="/xl/externalLinks/externalLink%s.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"/>' % n, '')

entries['xl/_rels/workbook.xml.rels'] = rels.encode('utf8')
entries['xl/workbook.xml'] = wbx.encode('utf8')
entries['[Content_Types].xml'] = ct.encode('utf8')

# externalLink parts need their own _rels (lost). Give each a stub rel.
for n in sorted(present):
    p = 'xl/externalLinks/_rels/externalLink%d.xml.rels' % n
    if p not in entries:
        entries[p] = (b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                      b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                      b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath"'
                      b' Target="UNKNOWN.xlsx" TargetMode="External"/></Relationships>')

if os.path.exists(OUT):
    os.remove(OUT)
with zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED) as z:
    for k, v in entries.items():
        z.writestr(k, v)
print("\nWROTE", OUT, os.path.getsize(OUT))

import openpyxl
wb = openpyxl.load_workbook(OUT, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"   '{ws.title}' state={ws.sheet_state} max_row={ws.max_row} max_col={ws.max_column}")
wb.close()
