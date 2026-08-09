import openpyxl, collections, re
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a): lines.append(' '.join(str(x) for x in a))

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# ---------- Highfield sheet ----------
ws = wb['Highfield']
hf = [r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20, values_only=True)]
p('='*100); p('HIGHFIELD SHEET: rows', len(hf))
cats = collections.Counter()
colours = collections.Counter()
for i, r in enumerate(hf[7:], start=8):
    if r[2] is None: continue
    cats[str(r[5])] += 1
    if r[4] is not None: colours[str(r[4])] += 1
p('Highfield col F "Category" DISTINCT (%d):' % len(cats))
for k, v in cats.most_common():
    p(f'   {v:5}  {k!r}')
p('Highfield col E "Colour" DISTINCT (%d):' % len(colours))
for k, v in colours.most_common(40):
    p(f'   {v:5}  {k!r}')

p('')
p('--- HIGHFIELD ROWS MATCHING 560 / SP560 / OS560 ---')
n = 0
for i, r in enumerate(hf, start=1):
    joined = ' | '.join('' if v is None else str(v) for v in r[:16])
    if re.search(r'560', joined):
        n += 1
        if n <= 80:
            p(f'  R{i}: Code={r[2]} | Equip={str(r[3])[:70]} | Colour={r[4]} | Cat={r[5]} | Price={r[6]} | AUD={r[7]} | Trade={r[11]} | Retail={r[15]}')
p('  total 560 matches in Highfield sheet:', n)

# ---------- Parts Maintenance rows mentioning 560 ----------
ws = wb['Parts Maintenance']
pm = [r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=30, values_only=True)]
p('')
p('='*100)
p('--- PARTS MAINTENANCE ROWS MENTIONING "560" ---')
n = 0
for i, r in enumerate(pm, start=1):
    j = ' '.join('' if v is None else str(v) for v in r[:8])
    if '560' in j:
        n += 1
        if n <= 60:
            p(f'  R{i}: C={str(r[2])[:75]} | D={r[3]} | E={r[4]} | Sell(L)={r[11]} | Install(N)={str(r[13])[:45]} | Y={r[24]}')
p('  total:', n)

# Highfield inflatables band in Parts Maintenance R2372..
p('')
p('--- PARTS MAINTENANCE: HIGHFIELD INFLATABLES BAND sample (R2372-2420) ---')
for i in range(2372, 2421):
    r = pm[i-1]
    if r[2] is None: continue
    p(f'  R{i}: C={str(r[2])[:70]} | D={r[3]} | E={r[4]} | G(P&A)={r[6]} | L(Sell)={r[11]} | N={str(r[13])[:40]} | Y={r[24]}')

# ---------- Dealer Fit HIGHFIELD - Sport 560 band ----------
ws = wb['Dealer Fit Module']
df = [r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=300, values_only=True)]
SLOT_START = [20 + 9*i for i in range(30)]
HDR = df[10]
p('')
p('='*100)
p('--- DEALER FIT MODULE: HIGHFIELD - Sport 560 BAND (rows 1598-1616) ---')
for i in range(1598, 1617):
    r = df[i-1]
    if r[2] is None: continue
    main = []
    for c in range(3, 20):
        v = r[c-1]
        if v is None or str(v).strip()=='': continue
        main.append(f'{HDR[c-1]}={str(v)[:60]}')
    p(f'  ROW {i}  ' + ' ; '.join(main))
    for si, s in enumerate(SLOT_START, 1):
        acc = r[s-1]
        if acc is None or str(acc).strip() in ('', '0', '.'): continue
        f = ['Accessory','Code','CTD','Sell','LabourOp','LabHrs','Sundry','Sublet']
        parts = [f'{f[k]}={str(r[s-1+k])[:60]}' for k in range(8) if r[s-1+k] is not None and str(r[s-1+k]).strip()!='']
        p(f'      slot{si}({get_column_letter(s)}): ' + ' ; '.join(parts))
    if r[288] is not None: p(f'      LongDesc[KE]={str(r[288])[:130]}')

# search all Dealer Fit rows mentioning 560
p('')
p('--- DEALER FIT MODULE: ALL ROWS whose Description/Code mentions 560 ---')
n = 0
for i, r in enumerate(df, start=1):
    j = f'{r[2]} {r[3]}'
    if '560' in j:
        n += 1
        if n <= 60:
            p(f'  R{i}: C={str(r[2])[:75]} | Code={r[3]} | ActSell(R)={r[17]} | ActCTD(O)={r[14]}')
p('  total:', n)

wb.close()
open(OUT+'p6_highfield_560.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
