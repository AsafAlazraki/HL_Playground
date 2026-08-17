"""Dump the Dealer Fit Module sheet into extracts/pd_dealerfit.json. READ-ONLY.

WHY THIS SHEET EXISTS AS ITS OWN EXTRACT, and is not folded into p1_parts.json:
FITMENT_RULES.md R3. The boat row's `Additional Dealer Fit Options - Line 01..42`
band (Boat Module!OL..QA) resolves into THIS sheet's column C at 99.4% live, and
into `Parts Maintenance`!C at 38.8%. They are two different libraries and merging
them would mis-resolve three pairs in five. 37 cells in Boat Module!OM are
literally `='[3]Dealer Fit Module'!$C$<row>`, which is the assertion.

SHAPE, established by probes/p5_dealerfit.py and re-confirmed here:
  row 11   the header row ("Description", "Code", "CTD", ... "Image Link")
  row 12+  data. Column C is the display name and the join key; column D is the
           supplier code. A row with C filled and D empty is a CATEGORY BANNER
           ("MINN KOTA MOTOR OPTIONS", "TRAILER SETUPS", "DEC RIGGING KITS"),
           exactly the convention Parts Maintenance uses — the category is a
           ROW, never a field.

Columns A..AB are kept: the identity and pricing ladder (C..S), the primary
bundled accessory (T..AA) and the image link (AB). The three further accessory
slots (AC..) and the 30-slot accessory grid beyond are not seeded.
"""
import openpyxl, json, sys
from pathlib import Path
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"

MAXC = ci("AB")

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", len(wb.sheetnames), file=sys.stderr)
ws = wb["Dealer Fit Module"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=2680, min_col=1, max_col=MAXC), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is not None and str(v).strip() != "":
            d[gl(i)] = v if isinstance(v, (int, float, bool)) else str(v)
    if d:
        data[r] = d

json.dump(data, open(OUT + "pd_dealerfit.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("dealer fit rows", len(data), file=sys.stderr)
wb.close()
