import openpyxl, sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"

# non-read-only needed for styles, but Dropdowns claims 1048576 rows; use read_only for values first
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Dropdowns"]
cols = {}
n = 0
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=400, max_col=20, values_only=True), 1):
    for j, v in enumerate(r):
        if v is not None and str(v).strip() != '':
            cols.setdefault(j, []).append((i, str(v).strip()))
wb.close()
from openpyxl.utils import get_column_letter as gcl
for j in sorted(cols):
    vals = cols[j]
    print("="*90)
    print(f"COLUMN {gcl(j+1)}  entries={len(vals)}  header={vals[0][1]}")
    for i, v in vals:
        print(f"   r{i:<4} {v}")
json.dump({gcl(j+1): cols[j] for j in cols}, open('t06_dropdowns.json','w',encoding='utf-8'), indent=0)
