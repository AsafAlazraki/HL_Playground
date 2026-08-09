import openpyxl, json, sys, io
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
lines = []
def p(*a):
    s = ' '.join(str(x) for x in a)
    lines.append(s)
    print(s)

p('SHEETS:', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    p(f'  {name!r:35} dims={ws.calculate_dimension()} rows={ws.max_row} cols={ws.max_column} state={ws.sheet_state}')

# print first 12 rows x 30 cols of each sheet, truncated
for name in wb.sheetnames:
    ws = wb[name]
    p('')
    p('='*100)
    p('SHEET:', name)
    p('='*100)
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=32, values_only=True), 1):
        cells = []
        for c, v in enumerate(row, 1):
            if v is not None and str(v).strip() != '':
                sv = str(v).replace('\n', '\\n')
                if len(sv) > 45: sv = sv[:45] + '...'
                cells.append(f'{get_column_letter(c)}{r}={sv}')
        if cells:
            p('  ' + ' | '.join(cells))
wb.close()

with open(OUT + 'p1_inventory.txt', 'w', encoding='utf8') as f:
    f.write('\n'.join(lines))
print('WROTE', OUT + 'p1_inventory.txt')
