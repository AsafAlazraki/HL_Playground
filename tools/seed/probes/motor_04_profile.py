import openpyxl, sys, json, collections, re
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

wb = openpyxl.load_workbook(SAL, read_only=True, data_only=True)
ws = wb['Motor Library']
data = []
for r in ws.iter_rows(min_row=1, max_row=2186, max_col=651, values_only=True):
    data.append(list(r) + [None]*(651-len(r)))
wb.close()
print("rows loaded", len(data))
hdr = data[3]  # row 4

def col(letter):
    from openpyxl.utils import column_index_from_string as ci
    return ci(letter)-1

# --- row structure: identify banner rows (C filled, D empty) vs data rows
banners = []
datarows = []
blank = 0
for i in range(4, len(data)):
    r = data[i]
    c, d = r[col('C')], r[col('D')]
    if c is not None and (d is None or str(d).strip() == ''):
        banners.append((i+1, c))
    elif d is not None and str(d).strip() != '':
        datarows.append(i+1)
    else:
        blank += 1
print("\nBANNER ROWS (row, label):", len(banners))
for rr, c in banners:
    print(f"   r{rr}: {c}")
print("DATA ROWS:", len(datarows), "first", datarows[:5], "last", datarows[-5:])
print("blank rows:", blank)

out = {'banners': banners, 'n_datarows': len(datarows)}

# --- profile key identity columns over data rows
KEY = ['C','D','E','F','G','H','I','J','K','L','N','O','P','Q','TZ','FF','FG','FH','FI','FA','FC','EY','EZ','FB','CI','CS','AC','CJ','Y','BE','BK','BR','BX']
print("\n=== KEY COLUMN PROFILES (data rows only) ===")
prof = {}
for L in KEY:
    ci_ = col(L)
    vals = [data[i-1][ci_] for i in datarows]
    nn = [v for v in vals if v is not None and str(v).strip() != '']
    cnt = collections.Counter(str(v) for v in nn)
    prof[L] = {'header': hdr[ci_], 'fill': len(nn), 'of': len(datarows), 'card': len(cnt),
               'top': cnt.most_common(60)}
    print(f"\n[{L}] '{hdr[ci_]}' fill={len(nn)}/{len(datarows)} card={len(cnt)}")
    if len(cnt) <= 60:
        for v, n in cnt.most_common():
            print(f"      {n:>5}  {v[:90]}")
    else:
        for v, n in cnt.most_common(25):
            print(f"      {n:>5}  {v[:90]}")
        print("      ... (%d distinct)" % len(cnt))

json.dump({k: {kk: (vv if kk != 'top' else vv) for kk, vv in v.items()} for k, v in prof.items()},
          open('motor_keycols.json', 'w'), default=str, indent=1)

# --- brand / supplier x model breakdown
sup = col('Q'); mod = col('C')
bybrand = collections.Counter()
for i in datarows:
    bybrand[str(data[i-1][sup])] += 1
print("\nSUPPLIER counts:", bybrand.most_common())

# brand prefix from column C ("Yamaha - F2.5SMHB")
pref = collections.Counter()
for i in datarows:
    v = data[i-1][mod]
    if v: pref[str(v).split(' - ')[0]] += 1
print("BRAND PREFIX (col C):", pref.most_common())
json.dump(datarows, open('motor_datarows.json', 'w'))
