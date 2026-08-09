import openpyxl
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a):
    s = ' '.join(str(x) for x in a); lines.append(s)

# NOT read_only, so merged_cells available -- but load only needed sheets? openpyxl has no partial load.
# Use read_only=False but data_only=True; 38MB, may be slow but fine once.
wb = openpyxl.load_workbook(SRC, read_only=False, data_only=True)

for name in ['Parts Enquiry', 'Parts Maintenance', 'Dealer Fit Module', 'Dropdowns', 'Highfield']:
    ws = wb[name]
    p('='*110)
    p('SHEET:', name, 'rows=', ws.max_row, 'cols=', ws.max_column)
    p('MERGED RANGES (%d):' % len(ws.merged_cells.ranges))
    mr = sorted([str(r) for r in ws.merged_cells.ranges])
    for r in mr[:200]:
        cr = ws.merged_cells.ranges
        # get value of top-left
        from openpyxl.utils.cell import range_boundaries
        c1, r1, c2, r2 = range_boundaries(r)
        v = ws.cell(row=r1, column=c1).value
        p(f'   {r:16} = {str(v)[:70] if v is not None else ""}')
    if len(mr) > 200:
        p('   ... %d more' % (len(mr)-200))

    # full header rows (first 12 rows) all columns
    p('HEADER GRID rows 1-12, all cols:')
    for r in range(1, 13):
        cells = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != '':
                sv = str(v).replace('\n', '\\n')
                if len(sv) > 42: sv = sv[:42]+'..'
                cells.append(f'{get_column_letter(c)}={sv}')
        if cells:
            p(f'  R{r}: ' + ' | '.join(cells))
wb.close()
open(OUT+'p2_core_headers.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
