import openpyxl, sys, json, collections
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
COLS = {ci(c):c for c in ["C","E","IM","QR","QT","QV","QX","QZ","RB"]}
wb=openpyxl.load_workbook(P, read_only=True, data_only=False); ws=wb["Boat Module"]
stat=collections.defaultdict(lambda: collections.Counter())
sample={}
for i,row in enumerate(ws.iter_rows(min_row=1,max_row=2310,values_only=True), start=1):
    E=row[4] if len(row)>4 else None
    if not E or i in (1,2,3,143,200,226,233,248,262,278,949,951,955): continue
    blk = "LIVE" if i<1005 else "OBSOLETE"
    for j,c in COLS.items():
        if c in ("C","E"): continue
        v=row[j-1] if len(row)>=j else None
        kind = "formula" if isinstance(v,str) and v.startswith("=") else ("hardcoded" if v is not None else "empty")
        stat[(blk,str(E),c)][kind]+=1
wb.close()
print(f"{'block':9} {'matrix':24} {'col':4} formula/hardcoded/empty")
for k in sorted(stat):
    c=stat[k]
    print(f"{k[0]:9} {k[1][:24]:24} {k[2]:4} f={c['formula']:4} h={c['hardcoded']:4} e={c['empty']:4}")
