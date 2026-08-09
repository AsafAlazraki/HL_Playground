import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci
SRC='MotorModule_salvaged.xlsx'
COLS=("C D E F G H I J K L N O P Q R S T U V W X Y Z AA AC AD AE AF AG AH AI AJ AK AT AU AV AX AY AZ "
      "BB BC BD BE BF BH BI BJ BK BL BO BP BQ BR BS BU BV BW BX BY "
      "CI CJ CK CL CQ CR CS DA DG EY EZ FA FB FC FF FG FH FI GT GZ TZ").split()
IDX={ci(c):c for c in COLS}; MAXC=max(IDX)
wb=load_workbook(SRC, read_only=True, data_only=True)
ws=wb['Motor Library']
out={}
rn=0
for row in ws.iter_rows(min_row=1,max_row=300,max_col=MAXC,values_only=True):
    rn+=1
    c=row[2] if len(row)>2 else None
    if not isinstance(c,str): continue
    if not any(k in c for k in ('F90XB','F115XB','F70','F150XB')): continue
    cells={}
    for i,v in enumerate(row,start=1):
        L=IDX.get(i)
        if L and v not in (None,''): cells[L]=v
    out[rn]=cells
wb.close()
json.dump(out,open('nm_motors.json','w',encoding='utf-8'),indent=1,default=str)
for r in sorted(out,key=int):
    print(r,'|',out[r].get('C'),'|',out[r].get('D'),'|E',out[r].get('E'),'|F',out[r].get('F'),'|H',out[r].get('H'),'|BF',out[r].get('BF'),'|BL',out[r].get('BL'))
