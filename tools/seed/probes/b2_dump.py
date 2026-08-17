"""Dump the Boat Module grid into extracts/b2_*.json. READ-ONLY on the workbook.

Writes into ../extracts, which is committed — stage two runs for a collaborator
who does not have the workbooks at all. It used to write into a machine-specific
temp directory, which meant the only copy of the seed's input lived somewhere
that gets cleaned up; gen_all.py was fixed the same way and this follows it.
"""
import openpyxl, json, sys
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"

# OK..QA is the dealer-fit band: OK "Additional Package Options" (one distinct
# value, and it is the header text) and OL..QA "Additional Dealer Fit Options -
# Line 01..42". FITMENT_RULES.md R3 resolves that band into the Dealer Fit
# Module sheet at 99.4%, which is the largest relationship the seed did not
# carry; without these columns the join cannot be built at all.
KEEP = []
for a, b in [("A", "U"), ("II", "LD"), ("LF", "OI"), ("OK", "QA"), ("QC", "RK"), ("SV", "UJ")]:
    KEEP += list(range(ci(a), ci(b) + 1))
KEEPSET = set(KEEP)
MAXC = max(KEEP)

HEADERS = [1, 2, 3, 143, 200, 226, 233, 248, 262, 278, 280, 955]

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Boat Module"]

data = {}
hdr = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=1010, min_col=1, max_col=MAXC), start=1):
    if r in HEADERS:
        d = {}
        for i, c in enumerate(row, start=1):
            v = c.value
            if v is not None and str(v).strip() != "":
                d[get_column_letter(i)] = str(v)
        hdr[r] = d
    d = {}
    for i, c in enumerate(row, start=1):
        if i not in KEEPSET:
            continue
        v = c.value
        if v is not None and str(v).strip() != "":
            d[get_column_letter(i)] = v if isinstance(v, (int, float, bool)) else str(v)
    if d:
        data[r] = d

json.dump(hdr, open(OUT + "b2_headers.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
json.dump(data, open(OUT + "b2_data.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("headers", len(hdr), "datarows", len(data), file=sys.stderr)
wb.close()
