import openpyxl, collections, re
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a): lines.append(' '.join(str(x) for x in a))

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb['Dealer Fit Module']
rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=300, values_only=True):
    rows.append(row)
wb.close()
p('rows read', len(rows))

# accessory slot start columns (1-based): T=20, then +9 each
SLOT_START = [20 + 9*i for i in range(30)]
HDR = rows[10]  # row 11

def val(r, c):
    v = r[c-1] if c-1 < len(r) else None
    return v

data = []
for i, r in enumerate(rows[11:], start=12):
    c = val(r, 3)
    if c is None or str(c).strip() == '':
        continue
    data.append((i, r))
p('rows with description in C:', len(data))

# slot usage histogram
slotcount = collections.Counter()
for i, r in data:
    n = 0
    for s in SLOT_START:
        v = val(r, s)
        if v is not None and str(v).strip() not in ('', '0'):
            n += 1
    slotcount[n] += 1
p('SLOT USAGE (# accessory lines populated per row):', dict(sorted(slotcount.items())))
maxslots = max(k for k in slotcount if k > 0)
p('max slots used =', maxslots)

# column profile main block C..S and KD..KF
p('')
p('--- MAIN BLOCK COLUMN PROFILE (cols C..S, KD..KF) ---')
for c in list(range(3, 20)) + [288, 289, 290]:
    vals = [val(r, c) for _, r in data]
    nz = [v for v in vals if v is not None and str(v).strip() != '']
    card = len(set(str(v) for v in nz))
    types = collections.Counter(type(v).__name__ for v in nz)
    h = HDR[c-1] if c-1 < len(HDR) else None
    p(f'  {get_column_letter(c):3} {str(h)[:26]:28} fill={len(nz)}/{len(data)} card={card} types={dict(types)}')
    if 0 < card <= 25:
        cnt = collections.Counter(str(v) for v in nz)
        p('       DISTINCT: ' + ' || '.join(f'{k[:50]!r}x{v}' for k, v in cnt.most_common()))

# section detection: rows where C set but D (code) empty AND E empty -> band header
p('')
p('--- POSSIBLE SECTION / BAND HEADER ROWS (C set, D and E empty) ---')
for i, r in data:
    if (val(r,4) in (None,'')) and (val(r,5) in (None,'')) and (val(r,20) in (None,'','0')):
        p(f'  R{i}: {str(val(r,3))[:110]}')

# code prefix distribution
p('')
p('--- CODE (col D) PREFIX DISTRIBUTION ---')
pref = collections.Counter()
for i, r in data:
    d = val(r, 4)
    if d is None: continue
    s = str(d)
    m = re.match(r'^([A-Za-z]+[-_ ]?)', s)
    pref[m.group(1) if m else s[:3]] += 1
for k, v in pref.most_common(50):
    p(f'   {v:5}  {k!r}')

# sample rows across bands with slots expanded
p('')
p('--- SAMPLE ROWS (main block + accessory slots) ---')
SLOT_FIELDS = ['Accessory','Code','CTD','Sell','Labour','LabHrs','Sundry','Sublet']
def dump(i, r, maxslot=8):
    main = []
    for c in range(3, 20):
        v = val(r, c)
        if v is None or str(v).strip()=='' : continue
        h = HDR[c-1]
        sv = str(v)
        if len(sv) > 55: sv = sv[:55]+'..'
        main.append(f'{h}={sv}')
    p(f'  ROW {i}  ' + ' ; '.join(main))
    for si, s in enumerate(SLOT_START[:maxslot], 1):
        acc = val(r, s)
        if acc is None or str(acc).strip() in ('', '0'): continue
        parts = []
        for k, f in enumerate(SLOT_FIELDS):
            v = val(r, s+k)
            if v is None or str(v).strip()=='': continue
            sv = str(v)
            if len(sv) > 55: sv = sv[:55]+'..'
            parts.append(f'{f}={sv}')
        p(f'      slot{si} ({get_column_letter(s)}): ' + ' ; '.join(parts))
    kd = val(r, 289)
    if kd: p(f'      LongDesc[KE]={str(kd)[:120]}')

for k in [0, 5, 40, 200, 400, 600, 700, 760, 900, 1200, 1600, 2000, len(data)-1]:
    if k < len(data):
        dump(*data[k])

open(OUT+'p5_dealerfit.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
