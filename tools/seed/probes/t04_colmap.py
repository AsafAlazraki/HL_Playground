import openpyxl, sys, io
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Trailer Module"]
rows = []
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=7, max_col=385, values_only=True), 1):
    rows.append(r)
h1 = rows[0]; num = rows[2]; d1 = rows[4]; d2 = rows[6]
print("COL | HDR(row1) | NUM(row3) | ROW5 | ROW7")
for j in range(385):
    a = h1[j] if j < len(h1) else None
    b = num[j] if j < len(num) else None
    c = d1[j] if j < len(d1) else None
    d = d2[j] if j < len(d2) else None
    if all(x is None for x in (a, b, c, d)):
        continue
    print(f"{gcl(j+1):>4} | {str(a)[:42]:<42} | {str(b)[:4]:<4} | {str(c)[:32]:<32} | {str(d)[:32]}")
wb.close()
