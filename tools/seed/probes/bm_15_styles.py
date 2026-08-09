import openpyxl, sys, json, collections
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
H=json.load(open(SP+"bm_headers.json")); hdr={gl(int(j)):v for j,v in H["278"]}; hdr1={gl(int(j)):v for j,v in H["1"]}
HH=lambda c: hdr.get(c) or hdr1.get(c,'')
wb=openpyxl.load_workbook(P, read_only=True, data_only=True)
ws=wb["Boat Module"]
TARGET={278,829,1534}
got={}
for i,row in enumerate(ws.iter_rows(min_row=1,max_row=1540), start=1):
    if i in TARGET:
        d={}
        for _j,c in enumerate(row,1):
            if not hasattr(c,'column'): continue
            if c.value is None and i!=278: continue
            try:
                fill=c.fill; fg = fill.fgColor.rgb if fill and fill.fgColor else None
                pat = fill.patternType if fill else None
                fnt=c.font; bold=fnt.bold; fc=fnt.color.rgb if fnt and fnt.color else None
                nf=c.number_format
            except Exception as e:
                fg=pat=bold=fc=nf=None
            d[gl(c.column)]=(pat, str(fg), bold, str(fc), nf)
        got[i]=d
wb.close()

for i in sorted(got):
    d=got[i]
    print(f"\n===== ROW {i} style runs =====")
    prev=None; start=None; keys=list(d.keys())
    for k in keys:
        sig=d[k]
        if sig!=prev:
            if prev is not None:
                print(f"  {start}..{last} : pat={prev[0]} fill={prev[1]} bold={prev[2]} fontcol={prev[3]} numfmt={prev[4]}  [{HH(start)} .. {HH(last)}]")
            prev=sig; start=k
        last=k
    if prev is not None:
        print(f"  {start}..{last} : pat={prev[0]} fill={prev[1]} bold={prev[2]} fontcol={prev[3]} numfmt={prev[4]}  [{HH(start)} .. {HH(last)}]")
