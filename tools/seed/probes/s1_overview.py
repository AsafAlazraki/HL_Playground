import openpyxl, json, sys
from openpyxl.utils import get_column_letter
SRC = r"C:/Users/AsafA/Downloads/MASTER PRICE FILE.xlsx"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
print("=== SHEETNAMES ===")
for n in wb.sheetnames:
    ws = wb[n]
    print(f"  {n!r:26} dims={ws.calculate_dimension()} max_row={ws.max_row} max_col={ws.max_column}")
wb.close()

# non-readonly pass for merged cells / validations / defined names
wb2 = openpyxl.load_workbook(SRC, data_only=True)
print("\n=== DEFINED NAMES ===")
try:
    for k,v in wb2.defined_names.items():
        print(f"  {k} -> {v.value}")
except Exception as e:
    print("  err", e)

print("\n=== MERGED CELLS PER SHEET ===")
for n in wb2.sheetnames:
    ws = wb2[n]
    mc = sorted([str(r) for r in ws.merged_cells.ranges])
    print(f"  [{n}] count={len(mc)}")
    for r in mc[:80]:
        print("     ", r)
    if len(mc)>80: print("      ...", len(mc)-80, "more")

print("\n=== DATA VALIDATIONS PER SHEET ===")
for n in wb2.sheetnames:
    ws = wb2[n]
    dvs = list(ws.data_validations.dataValidation)
    print(f"  [{n}] count={len(dvs)}")
    for dv in dvs:
        print(f"     type={dv.type} sqref={dv.sqref} formula1={dv.formula1!r}")
wb2.close()
