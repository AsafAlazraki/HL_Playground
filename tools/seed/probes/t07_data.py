import openpyxl, sys, io, json
from collections import Counter, OrderedDict
from openpyxl.utils import get_column_letter as gcl, column_index_from_string as cifs
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"

BRAND_BANNERS = {186:'GFAB TRAILERS',232:'STACER TRAILERS',281:'DUNBIER TRAILERS',
                 455:'MACKAY TRAILERS',626:'DUNBIER / HAINES BMT TRAILERS',656:'OBSOLETE'}

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Trailer Module"]
rows = [r for r in ws.iter_rows(min_row=1, max_row=918, max_col=385, values_only=True)]
wb.close()

def G(r, L):
    j = cifs(L)-1
    return r[j] if j < len(r) else None
def ne(v):
    return v is not None and str(v).strip() != '' and str(v).strip() != '.'

SPEC = OrderedDict([('C','Brand / Make / Model'),('D','Supplier'),('E','Code'),('F','Long Description'),
    ('G','Image Link'),('H','Boat Size (Mtr)'),('I','Wheel Size'),('J','Tare (Kg)'),('K','ATM (KG)'),
    ('L','Winch'),('M','Between Guards (mm)'),('N','Trailer Length (Mtr)'),('O','Trailer Plug'),
    ('AN','Dealer'),('AO','Discount'),('AP','Settlement'),('AQ','Nett Price'),('AR','Freight'),('AS','Landed'),
    ('AV','Lockout Date'),('AW','Factory Build Date'),('AX','Factory Completion Date'),('AY','Factory Shipping Date'),('AZ','Estimated Lead Time (Days)'),
    ('BB','PD Operation'),('BC','PD (Hrs)'),('BD','PD ($)'),('BO','Sundry'),('BP','Detailing'),('BQ','Total PD Charges'),
    ('BS','Total Nett CTD'),('BT','MU %'),('BU','GP'),('BV','RRP'),('BW','Sell'),
    ('BY','Rego Type'),('BZ','Rego ($)'),('CA','Sell inc Rego')])

# classify rows
brand = None; series = None; recs = []
for i, r in enumerate(rows, 1):
    if i in BRAND_BANNERS:
        brand = BRAND_BANNERS[i]; series = None; continue
    a = G(r,'A'); c = G(r,'C'); e = G(r,'E')
    cnt = sum(1 for v in r if ne(v))
    if ne(a) and cnt <= 4:
        # series row; if no brand banner seen yet -> REDCO/TINKA or NSM CUSTOM
        series = str(c).strip() if ne(c) else str(a).strip()
        if brand is None: brand = 'REDCO / TINKA TRAILERS'
        continue
    if ne(c) and not ne(e) and cnt <= 5:
        series = str(c).strip(); continue
    if ne(c) and ne(e):
        d = {'row': i, 'brand': brand, 'series': series}
        for L, h in SPEC.items():
            d[L] = G(r, L)
        # options
        fo = []
        start = cifs('CD')
        for k in range(20):
            b = start + k*4
            code = r[b-1] if b-1 < len(r) else None
            desc = r[b]   if b   < len(r) else None
            cost = r[b+1] if b+1 < len(r) else None
            sell = r[b+2] if b+2 < len(r) else None
            if ne(code) or ne(desc):
                fo.append({'n':k+1,'code':code,'desc':desc,'cost':cost,'sell':sell})
        d['factory_options'] = fo
        dfo = []
        for k in range(20):
            b = cifs('FZ') + k
            v = r[b-1] if b-1 < len(r) else None
            if ne(v): dfo.append(str(v).strip())
        d['dealer_fit_options'] = dfo
        feats = []
        for k in range(21):
            b = cifs('R') + k
            v = r[b-1] if b-1 < len(r) else None
            if ne(v): feats.append(str(v).strip())
        d['features'] = feats
        recs.append(d)

print("DATA ROWS:", len(recs))
agg = OrderedDict()
for d in recs:
    agg.setdefault(d['brand'], OrderedDict()).setdefault(d['series'], []).append(d)
print("\n=== BRAND > SERIES > COUNT ===")
for b, sm in agg.items():
    print(f"BRAND {b}  series={len(sm)} trailers={sum(len(v) for v in sm.values())}")
    for s, v in sm.items():
        print(f"    {s} :: {len(v)}  rows {v[0]['row']}-{v[-1]['row']}")

print("\n=== COLUMN FILL / CARDINALITY ===")
for L, h in SPEC.items():
    vals = [d[L] for d in recs]
    nn = [v for v in vals if ne(v)]
    card = len(set(str(v) for v in nn))
    print(f"{L:<3} {h:<28} fill={len(nn):>3}/{len(recs)} card={card}")
    if card <= 30 and card > 0:
        cc = Counter(str(v) for v in nn).most_common()
        print("      ENUM: " + " | ".join(f"{k}({n})" for k, n in cc))

json.dump(recs, open('t07_records.json','w',encoding='utf-8'), indent=1, default=str)
print("\nsaved t07_records.json")
