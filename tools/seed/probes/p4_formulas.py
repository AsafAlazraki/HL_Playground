import openpyxl, re, collections
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a): lines.append(' '.join(str(x) for x in a))

wb = openpyxl.load_workbook(SRC, read_only=False, data_only=False)

TARGETS = ['Parts Enquiry', 'Parts Maintenance', 'Dealer Fit Module', 'Highfield', 'Dropdowns']

def norm(f):
    # strip row numbers -> {r}
    f = re.sub(r'(\$?[A-Z]{1,3}\$?)(\d+)', lambda m: m.group(1) + '{r}', f)
    return f

for name in TARGETS:
    ws = wb[name]
    p('='*110)
    p('SHEET:', name)
    # data validations
    try:
        dvs = ws.data_validations.dataValidation
    except Exception:
        dvs = []
    p('DATA VALIDATIONS (%d):' % len(dvs))
    for dv in dvs[:60]:
        p(f'   type={dv.type} formula1={str(dv.formula1)[:160]} sqref={str(dv.sqref)[:200]}')
    # formulas
    percol = collections.defaultdict(collections.Counter)
    example = {}
    total = 0
    maxr = min(ws.max_row, 4000)
    for row in ws.iter_rows(min_row=1, max_row=maxr, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                total += 1
                n = norm(v)
                percol[cell.column][n] += 1
                key = (cell.column, n)
                if key not in example:
                    example[key] = cell.coordinate + ' :: ' + v[:220]
    p('TOTAL FORMULA CELLS (rows<=%d): %d' % (maxr, total))
    hdrrow = {'Parts Maintenance': 1, 'Dealer Fit Module': 11, 'Highfield': 6, 'Parts Enquiry': 6, 'Dropdowns': 2}[name]
    for col in sorted(percol):
        cnt = percol[col]
        h = ws.cell(row=hdrrow, column=col).value
        p(f'  COL {get_column_letter(col)} ({str(h)[:40]}) distinct_patterns={len(cnt)} cells={sum(cnt.values())}')
        for pat, n in cnt.most_common(5):
            p(f'      x{n:5}  {pat[:260]}')
            p(f'             eg {example[(col,pat)][:280]}')
wb.close()
open(OUT+'p4_formulas.txt','w',encoding='utf8').write('\n'.join(lines))
print('done', len(lines))
