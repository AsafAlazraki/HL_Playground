"""Record WHICH fan-out cells on a boat row are FORMULAS. READ-ONLY.

WHY. FITMENT_RULES.md §9 settles what `__origin` on a pair means, and it is not
a constant:

    "`__origin` already carries the distinction the workbook itself recorded —
     86 trailer cells are live external links and 1,109 are typed strings,
     which is 'rule' versus 'added' written in the difference between a formula
     and a literal."

A cell reading `='[7]Trailer Module'!$C$140` is the business POINTING at the
library row: the pairing is derived, and that is `origin: 'rule'`. A cell
holding the same text typed by hand is a person putting it there, which is
`origin: 'added'`. Every other extract in this directory is dumped with
`data_only=True` and therefore cannot tell the two apart — it sees the cached
value either way. This probe is the only one that opens the workbook with
`data_only=False`, and it keeps nothing but the addresses.

OUTPUT. `extracts/b3_formula.json` = { "<row>": "KZ LF NZ OL ..." } — for each
boat row, the space-joined column letters in the fan-out bands whose cell is a
formula. Rows with no formula cell are omitted. Values are deliberately NOT
stored: the formula text is provenance, not data, and the cached value is
already in b2_data.json.
"""
import openpyxl, json, sys, collections
from pathlib import Path
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"

# The four fan-out bands, and only those: the 13 motor slots (name column only),
# the 10 trailer slots, the 42 dealer-fit lines and the 10 P/D part lines.
MOTOR_NAME_COLS = ["KZ", "LF", "LL", "LR", "LX", "MD", "MJ", "MP", "MV", "NB", "NH", "NN", "NT"]
TRAILER_COLS = ["NZ", "OA", "OB", "OC", "OD", "OE", "OF", "OG", "OH", "OI"]
DEALER_COLS = [gl(i) for i in range(ci("OL"), ci("QA") + 1)]
PD_COLS = [gl(i) for i in range(ci("JT"), ci("KC") + 1)]

WANT = MOTOR_NAME_COLS + TRAILER_COLS + DEALER_COLS + PD_COLS
WANTIDX = {ci(L): L for L in WANT}
MAXC = max(WANTIDX)

wb = openpyxl.load_workbook(P, read_only=True, data_only=False)
ws = wb["Boat Module"]

out = {}
tally = collections.Counter()
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=1010, min_col=1, max_col=MAXC), start=1):
    hits = []
    for i, c in enumerate(row, start=1):
        L = WANTIDX.get(i)
        if L is None:
            continue
        v = c.value
        if isinstance(v, str) and v.startswith("="):
            hits.append(L)
            tally["formula"] += 1
        elif v is not None and str(v).strip() != "":
            tally["literal"] += 1
    if hits:
        out[r] = " ".join(hits)

json.dump(out, open(OUT + "b3_formula.json", "w", encoding="utf-8"), ensure_ascii=False)
print("rows with >=1 formula cell", len(out), dict(tally), file=sys.stderr)
wb.close()
