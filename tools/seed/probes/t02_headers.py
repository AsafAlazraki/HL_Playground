import openpyxl, sys, io
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)

def head(name, nrows, ncols):
    ws = wb[name]
    print("="*100)
    print("SHEET:", name)
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=nrows, max_col=ncols, values_only=True), 1):
        rows.append(r)
    for i, r in enumerate(rows, 1):
        cells = [(gcl(j+1), v) for j, v in enumerate(r) if v is not None and str(v).strip() != '']
        if not cells:
            print(f"-- row {i}: (empty)")
            continue
        print(f"-- row {i}: n={len(cells)}")
        print("   " + " | ".join(f"{c}={str(v)[:45]}" for c, v in cells[:60]))

head("Trailer Spec Enquiry", 12, 40)
head("Trailer Module", 8, 385)
head("Dropdowns", 12, 20)
wb.close()
