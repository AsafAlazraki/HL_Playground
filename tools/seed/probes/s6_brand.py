import openpyxl, sys
from openpyxl.utils import get_column_letter
SRC = r"C:/Users/AsafA/Downloads/MASTER PRICE FILE.xlsx"
wv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
wf = openpyxl.load_workbook(SRC, read_only=True, data_only=False)
BRANDS=["Stabicraft","Surtees","Stacer","Formosa","Jeanneau","Haines Signature","Highfield","Boat Show"]
which = sys.argv[1] if len(sys.argv)>1 else None
for b in BRANDS:
    if which and b!=which: continue
    ws=wv[b]; wsf=wf[b]
    mr, mc = ws.max_row, ws.max_column
    rowsv=list(ws.iter_rows(min_row=1,max_row=mr,max_col=mc,values_only=True))
    rowsf=list(wsf.iter_rows(min_row=1,max_row=mr,max_col=mc,values_only=True))
    print(f"\n{'='*100}\n### {b}  ({mr} x {mc})\n{'='*100}")
    N = int(sys.argv[2]) if len(sys.argv)>2 else 14
    for i,(rv,rf) in enumerate(zip(rowsv,rowsf), start=1):
        if i>N: break
        cells=[]
        for j,(v,f) in enumerate(zip(rv,rf), start=1):
            if v is None and f is None: continue
            L=get_column_letter(j); s=f"{L}{i}="
            if isinstance(f,str) and f.startswith("="):
                s+=f"[F]{f[:200]}"
                if v is not None: s+=" ->"+repr(v)[:60]
            else: s+=repr(v)[:150]
            cells.append(s)
        if cells: print(f" r{i}: " + "\n       | ".join(cells))
wv.close(); wf.close()
