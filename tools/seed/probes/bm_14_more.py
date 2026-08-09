import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
H=json.load(open(SP+"bm_headers.json"))
hdr={gl(int(j)):v for j,v in H["278"]}; hdr1={gl(int(j)):v for j,v in H["1"]}
HH=lambda c: hdr.get(c) or hdr1.get(c,'')

ROWS = set([280,1436,1617,1125,1188,1266,1005,1533] + list(range(1534,1549)) + list(range(829,845)))
res={}
for mode in (True,False):
    wb=openpyxl.load_workbook(P, read_only=True, data_only=mode)
    ws=wb["Boat Module"]
    d={}
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=1620,values_only=True), start=1):
        if i in ROWS:
            d[i]={gl(j):v for j,v in enumerate(row,1) if v is not None and str(v).strip()!=""}
    wb.close(); res["v" if mode else "f"]=d
json.dump({k:{str(r):{c:str(v) for c,v in cc.items()} for r,cc in dd.items()} for k,dd in res.items()},
          open(SP+"bm_sp560_raw2.json","w"), indent=1)
V=res["v"];F=res["f"]
for r in [1005,280,1125,1436,1533]:
    print(f"\n===== SECTION ROW {r} =====")
    for c,v in V.get(r,{}).items():
        f=F.get(r,{}).get(c); fs=f"  <<{f[:90]}>>" if isinstance(f,str) and f.startswith("=") else ""
        print(f"  {c}{r} [{HH(c)}] = {str(v)[:150]}{fs}")

print("\n\n===== SP560 rows: OL..QA dealer-fit, KJ, SV..UJ PD/labour, RN..SQ graphics =====")
COLS = ["KJ","OL","OM","ON","OO"] + [f"S{x}" for x in "VWXYZ"] + ["TA","TB","TC","TD","TE","TF","TH","TI","TJ","TK","TL","TM","TN","TO","TP","TQ","TR","TT","TU","TV","TW","TX","TY","TZ","UA","UB","UC","UD","UF","UG","UH","UI","UJ"]
for r in range(1534,1549):
    d=V.get(r,{})
    parts=[]
    for c in COLS:
        if c in d and str(d[c]) not in (".",):
            parts.append(f"{c}={str(d[c])[:60]}")
    print(f"\nR{r} {d.get('C')} ({d.get('D')}):")
    for p in parts: print("   ", p)
    if r==1534:
        print("   FORMULAS:", {c:F[1534][c][:80] for c in COLS if c in F.get(1534,{}) and str(F[1534][c]).startswith("=")})
        break

print("\n\n===== SP560 all variants: RN..SQ Paint & Graphics =====")
for r in range(1534,1549):
    d=V.get(r,{})
    g=[f"{c}:{d[c]}" for c in [gl(x) for x in range(ci('RM'),ci('ST')+1)] if c in d]
    print(f"R{r} {d.get('D')} {d.get('C')}: " + " | ".join(g))
