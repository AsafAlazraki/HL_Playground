import openpyxl, sys, io, re, json
from collections import Counter, OrderedDict
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"

wb = openpyxl.load_workbook(P, read_only=True, data_only=False)
for SHEET, MAXR, MAXC in [("Trailer Module", 918, 385), ("Trailer Spec Enquiry", 91, 40)]:
    ws = wb[SHEET]
    percol = OrderedDict()
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=MAXR, max_col=MAXC), 1):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                norm = re.sub(r'(?<=[A-Z$])\d+', '{r}', v)
                norm = re.sub(r'\{r\}(\{r\})+', '{r}', norm)
                percol.setdefault(cell.column_letter, Counter())[norm] += 1
    print("="*100)
    print("FORMULAS IN", SHEET, " columns with formulas:", len(percol))
    for L, cnt in percol.items():
        tot = sum(cnt.values())
        print(f"-- {L}  (n={tot}, distinct={len(cnt)})")
        for f, n in cnt.most_common(4):
            print(f"      {n:>4}x  {f[:200]}")
wb.close()
