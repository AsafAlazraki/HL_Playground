import zipfile, re, sys, json, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')
SAL = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/MotorModule_salvaged.xlsx"

z = zipfile.ZipFile(SAL)
x = z.read('xl/worksheets/sheet1.xml').decode('utf8')
# x14 dataValidations live in extLst
ext = re.findall(r'<x14:dataValidation\b.*?</x14:dataValidation>', x, re.S)
print("X14 DATAVALIDATIONS:", len(ext))
for d in ext[:80]:
    d2 = re.sub(r'\s+', ' ', d)
    f1 = re.search(r'<xm:f>([^<]*)</xm:f>', d2)
    sq = re.search(r'<xm:sqref>([^<]*)</xm:sqref>', d2)
    ty = re.search(r'type="([^"]+)"', d2)
    print(f"   sqref={sq.group(1) if sq else '?':<22} type={ty.group(1) if ty else '?':<8} src={f1.group(1) if f1 else '?'}")
z.close()

# ---- header grid rows 1..5 for all columns
wb = openpyxl.load_workbook(SAL, read_only=True, data_only=True)
ws = wb['Motor Library']
rows = []
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
    rows.append(r)
wb.close()
ncol = max(len(r) for r in rows)
print("\nNCOL:", ncol)
out = []
for c in range(ncol):
    L = get_column_letter(c+1)
    vals = [rows[i][c] if c < len(rows[i]) else None for i in range(6)]
    if any(v is not None for v in vals[:4]):
        out.append((L, vals))
print("cols with any header in rows1-4:", len(out))
with open('motor_header_grid.txt', 'w', encoding='utf8') as f:
    for L, vals in out:
        f.write(f"{L}\t" + "\t".join("" if v is None else str(v).replace("\n", "\\n") for v in vals) + "\n")
for L, vals in out[:400]:
    print(L, "|", " || ".join("" if v is None else str(v).replace("\n", " ")[:40] for v in vals))
