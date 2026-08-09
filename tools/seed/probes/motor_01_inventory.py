import openpyxl, json, sys, io
sys.stdout.reconfigure(encoding='utf-8')

PATH = r"C:/Users/AsafA/Downloads/Motor Module (1).xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("SHEETS:", len(wb.sheetnames))
for ws in wb.worksheets:
    print(f"  '{ws.title}' state={ws.sheet_state} dims={ws.calculate_dimension()} max_row={ws.max_row} max_col={ws.max_column}")
wb.close()

# also non-read-only for merged cells / defined names (may be slow, do metadata only)
wb2 = openpyxl.load_workbook(PATH, read_only=False, data_only=False)
print("\nDEFINED NAMES:")
try:
    for name, dn in wb2.defined_names.items():
        print("  ", name, "->", dn.value)
except Exception as e:
    print("  err", e)
print("\nMERGED RANGES per sheet:")
for ws in wb2.worksheets:
    mr = list(ws.merged_cells.ranges)
    print(f"  '{ws.title}': {len(mr)} merged")
    for r in mr[:40]:
        print("     ", str(r), "=", repr(ws.cell(r.min_row, r.min_col).value))
wb2.close()
