import openpyxl, sys, json, time
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl

P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
t0=time.time()
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
ws = wb["Boat Module"]

recs=[]
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    n=0; mx=0
    for j,v in enumerate(row, start=1):
        if v is not None and str(v).strip()!="":
            n+=1
            mx=j
    C = row[2] if len(row)>2 else None
    D = row[3] if len(row)>3 else None
    recs.append({"r":i,"n":n,"maxc":mx,"C":C,"D":D})
wb.close()
print("pass secs", round(time.time()-t0,1))
json.dump(recs, open(SP+"bm_rowmap.json","w"), indent=0, default=str)

# header rows: D == 'Model Code'
hdr=[r for r in recs if isinstance(r["D"],str) and r["D"].strip()=="Model Code"]
print("HEADER ROWS (D='Model Code'):", len(hdr))
for r in hdr:
    print(f"  R{r['r']}: C='{r['C']}'  nonempty={r['n']} maxcol={r['maxc']}({gl(r['maxc']) if r['maxc'] else '-'})")

print("\nMAX maxc overall:", max(r["maxc"] for r in recs), " row:", max(recs,key=lambda r:r["maxc"])["r"])
# rows with C set but D empty -> section/range headers
sec=[r for r in recs if r["C"] not in (None,"") and (r["D"] in (None,"")) ]
print("\nSECTION/RANGE HEADER ROWS (C set, D empty):", len(sec))
for r in sec[:400]:
    print(f"  R{r['r']}: '{r['C']}' n={r['n']} maxc={gl(r['maxc']) if r['maxc'] else '-'}")
