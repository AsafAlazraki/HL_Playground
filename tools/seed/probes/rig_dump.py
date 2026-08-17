"""Dump Rigging Module.xlsx · sheet "Rigging Kits" -> extracts/rig_kits.json. READ-ONLY.

THE EIGHTH WORKBOOK, and it is here now. `FITMENT_RULES.md` §6.5 and Appendix B
item 2 told the owner the embedded cache was "demonstrably incomplete" and asked
for this file to fill a hole. `FOUR_MODULES.md` §3.1 struck both sentences: the
cache is a byte-exact mirror (42,372 cells present in both, **0** differing, 0
missing either way) and the demonstration ran the wrong way round. What the real
file adds is the DERIVATION, not the data — the price ladder, the two external
labour rates, the `OBSOLETE RIGGING KITS` divider at `C829` and the six preamble
sentinels. We read it directly rather than through the cache because reading a
sheet is simpler than reading a mirror of one, and because the divider and the
banner rows are structure the cache flattens.

SHAPE, per FOUR_MODULES.md §3.5 / §3.6, re-confirmed here first-hand:

  row 1        the master header labels (C RIGGING KIT DESCRIPTION · D Part
               Number · E Build · F Dealer · … · BE Dealer 1/7/22)
  row 2        the RATE CONSTANTS that drive the ladder, not data:
               K2 = 0.25 (kit markup) · L2 = M2 = 0.05 · O2 = "(Hrs)" ·
               P2 = 130.09090909090907, which is `Service Module!Labour
               Rates!$G$14` — the cost rate — reaching in from another workbook.
  row 3        column ordinals 1..55, the workbook's own VLOOKUP index row
  rows 4-9     SIX PREAMBLE SENTINELS, and they are not inert. `C4 =
               "NR - RIGGING KIT NOT REQUIRED"`, but `C5 = "SUP - Supplied
               Standard w Motor"` carries **O5 = 2.5 hours**: the motor ships
               with the kit and the fitting is still billed. Dropping row 5
               silently deletes money from every quote that names it. They are
               imported as real rows, flagged, never dropped.
  rows 11-828  the live catalogue, opened by band header rows and separated by
               "." spacer rows.
  row 827      `NR - RIGGING KIT NOT REQUIRED` again, verbatim — the last live
               row, and the reason the importer tests the STRING and not the
               row number.
  row 829      `OBSOLETE RIGGING KITS` — the divider. Third instance of the same
               mechanism: `Boat Module!A1005`, `Trailer Module!A656`,
               `Parts Maintenance!C2918`, and this.
  rows 830+    the obsolete catalogue. 153 live triples on 31 live boat rows
               still point below this divider, NINE of them in slot 1 — the
               boat's standard fit (FOUR_MODULES.md §3.9).

A BAND HEADER is a row whose `D Part Number` is empty and whose `K` repeats the
header word "Sell Price" — the band's own re-labelling of the price columns
(`F` reads "Dealer", " Fact. Cost ", " Base Cost " or " EURO " depending on the
band, which is the only place the currency of `F` is written down). A "." row is
a spacer. Both are dumped verbatim; classifying them is stage two's job.

Columns A..BM, rows 1..1460. The sheet declares 3,039 rows and 65 columns; the
last populated row is 1,451 and everything past it is `\\xa0` padding.
"""
import openpyxl, json, sys
from pathlib import Path
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci

P = r"C:/Users/AsafA/Downloads/Rigging Module.xlsx"
OUT = str(Path(__file__).resolve().parent.parent / "extracts") + "/"
MAXC = ci("BM")
MAXR = 1460
NBSP = "\u00a0"

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("SHEETS", wb.sheetnames, file=sys.stderr)
ws = wb["Rigging Kits"]
print("DIMS", ws.max_row, ws.max_column, file=sys.stderr)

data = {}
for r, row in enumerate(ws.iter_rows(min_row=1, max_row=MAXR, min_col=1, max_col=MAXC), start=1):
    d = {}
    for i, c in enumerate(row, start=1):
        v = c.value
        if v is None:
            continue
        if isinstance(v, str):
            # the padding this sheet is mostly made of
            if v.replace(NBSP, " ").strip() == "":
                continue
            d[gl(i)] = v
        elif isinstance(v, (int, float, bool)):
            d[gl(i)] = v
        else:
            d[gl(i)] = str(v)
    if d:
        data[r] = d

json.dump(data, open(OUT + "rig_kits.json", "w", encoding="utf-8"), ensure_ascii=False, default=str)
print("rigging rows", len(data), "last", max(data), file=sys.stderr)
band = [r for r in sorted(data) if "C" in data[r] and str(data[r].get("K", "")).strip() == "Sell Price"]
solo = [r for r in sorted(data) if list(data[r]) == ["C"]]
print("band header rows", len(band), file=sys.stderr)
print("C-only rows (divider, sub-banners, '.' spacers)", solo, file=sys.stderr)
wb.close()
