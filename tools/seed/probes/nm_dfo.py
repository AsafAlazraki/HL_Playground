import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci
SRC=r"C:/Users/AsafA/Downloads/Parts Module (3).xlsx"
COLS="C D E F G H I J K L M N O P Q R S".split()
IDX={ci(c):c for c in COLS}; MAXC=max(IDX)
NEEDLES=['GX750B Hideaway','Apollo RA670','EchoMap Ultra 2 125sv','Tube Covers to suit PVC Boat - 5.6','Tube Covers to suit Hypalon Boat - 5.6']
wb=load_workbook(SRC, read_only=True, data_only=True)
ws=wb['Dealer Fit Module']
out={}; rn=0
for row in ws.iter_rows(min_row=1,max_row=2680,max_col=MAXC,values_only=True):
    rn+=1
    c=row[2] if len(row)>2 else None
    if not isinstance(c,str): continue
    if not any(n.lower() in c.lower() for n in NEEDLES): continue
    cells={}
    for i,v in enumerate(row,start=1):
        L=IDX.get(i)
        if L and v not in (None,''): cells[L]=v
    out[rn]=cells
wb.close()
json.dump(out,open('nm_dfo.json','w',encoding='utf-8'),indent=1,default=str)
for r in sorted(out,key=int):
    c=out[r]
    print(r,'|',c.get('C'),'| code',c.get('D'),'| CTD E',c.get('E'),'| PartsSell I',c.get('I'),'| Lab J',c.get('J'),'| ActCTD O',c.get('O'),'| MU P',c.get('P'),'| GP Q',c.get('Q'),'| ActSell R',c.get('R'))
