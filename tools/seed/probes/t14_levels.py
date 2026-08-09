import openpyxl, sys, io, json
from collections import Counter
from openpyxl.utils import get_column_letter as gcl
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
P = r"C:/Users/AsafA/Downloads/Trailer Module.xlsx"
wb = openpyxl.load_workbook(P, read_only=False, data_only=True)
ws = wb["Trailer Module"]
print("=== ROW LEVEL BY FONT SIZE (col C), rows 1-720 ===")
lvl = []
for r in range(1, 721):
    c = ws.cell(row=r, column=3)
    if c.value is None or str(c.value).strip() == '': continue
    e = ws.cell(row=r, column=5).value
    sz = c.font.size; bold = c.font.bold
    fill = c.fill.start_color.rgb if c.fill and c.fill.start_color else None
    fill = fill if isinstance(fill, str) else 'theme'
    has_code = e is not None and str(e).strip() not in ('', '.')
    lvl.append((r, sz, bold, fill, has_code, str(c.value)[:58]))
sizes = Counter((s, b, hc) for _, s, b, _, hc, _ in lvl)
print("  (size,bold,hasCode) ->", dict(sizes))
print("\n  --- all rows with size>=12 (structural) ---")
for r, s, b, f, hc, v in lvl:
    if s and s >= 12:
        print(f"   r{r:<4} size={s} bold={b} fill={f:<10} code={hc}  {v}")
wb.close()
