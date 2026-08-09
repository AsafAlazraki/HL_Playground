import openpyxl, collections, json
from openpyxl.utils import get_column_letter

SRC = r'C:/Users/AsafA/Downloads/Parts Module (3).xlsx'
OUT = r'C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/'
lines = []
def p(*a):
    lines.append(' '.join(str(x) for x in a))

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb['Parts Maintenance']
HDR = 1
rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=62, values_only=True):
    rows.append(row)
wb.close()
p('TOTAL ROWS READ', len(rows))
hdr = rows[0]
headers = {}
for c, v in enumerate(hdr, 1):
    if v is not None and str(v).strip():
        headers[c] = str(v).strip()
p('HEADERS:', json.dumps({get_column_letter(k): v for k, v in headers.items()}, ensure_ascii=False))

# classify rows
CAT_ROWS = []   # category header rows: col C set, col E (code) empty
DATA_ROWS = []
BLANK = 0
for i, r in enumerate(rows[3:], start=4):  # 1-indexed excel row
    cC = r[2]; cD = r[3]; cE = r[4]
    nonempty = sum(1 for v in r if v is not None and str(v).strip() != '')
    if nonempty == 0:
        BLANK += 1
        continue
    if cC is not None and (cE is None or str(cE).strip() == ''):
        CAT_ROWS.append((i, str(cC), str(cD) if cD else ''))
    elif cE is not None:
        DATA_ROWS.append((i, r))
p('CATEGORY HEADER ROWS:', len(CAT_ROWS), 'DATA ROWS:', len(DATA_ROWS), 'BLANK:', BLANK)
p('')
p('--- FULL CATEGORY LIST (verbatim, excel row) ---')
for i, c, d in CAT_ROWS:
    p(f'  R{i}: {c}')

# fill rate + cardinality per column
p('')
p('--- COLUMN PROFILE (over data rows) ---')
for c in range(1, 63):
    vals = [r[c-1] for _, r in DATA_ROWS]
    nz = [v for v in vals if v is not None and str(v).strip() != '']
    if not nz and c not in headers:
        continue
    types = collections.Counter(type(v).__name__ for v in nz)
    card = len(set(str(v) for v in nz))
    p(f'  {get_column_letter(c):3} {headers.get(c,"(no header)")[:32]:34} fill={len(nz)}/{len(DATA_ROWS)} card={card} types={dict(types)}')
    if 0 < card <= 40:
        cnt = collections.Counter(str(v) for v in nz)
        p('        DISTINCT: ' + ' || '.join(f'{k!r}x{v}' for k, v in cnt.most_common()))

# top values for medium-cardinality important cols
for c, nm in [(4,'Supplier'), (14,'Install Type')]:
    cnt = collections.Counter(str(r[c-1]) for _, r in DATA_ROWS if r[c-1] is not None)
    p('')
    p(f'--- TOP VALUES col {get_column_letter(c)} {nm} (card={len(cnt)}) ---')
    for k, v in cnt.most_common(60):
        p(f'   {v:5}  {k[:90]}')

# sample representative data rows
p('')
p('--- SAMPLE DATA ROWS ---')
import math
idxs = [0, 50, 200, 500, 900, 1400, 2000, 2600, 3000, len(DATA_ROWS)-1]
for k in idxs:
    if k >= len(DATA_ROWS): continue
    i, r = DATA_ROWS[k]
    kv = []
    for c in range(1, 63):
        v = r[c-1]
        if v is None or str(v).strip()=='': continue
        h = headers.get(c, get_column_letter(c))
        sv = str(v)
        if len(sv) > 60: sv = sv[:60]+'..'
        kv.append(f'{h}[{get_column_letter(c)}]={sv}')
    p(f'  ROW {i}: ' + ' ; '.join(kv))

open(OUT+'p3_parts_maint.txt','w',encoding='utf8').write('\n'.join(lines))
print('done lines', len(lines))
