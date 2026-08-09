import openpyxl, sys, json, collections
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.utils import get_column_letter as gl
P = r"C:/Users/AsafA/Downloads/Boat Module (5).xlsx"
SP = r"C:/Users/AsafA/AppData/Local/Temp/claude/C--Users-AsafA--claude-projects-HelmLogic-Dynamic-Config/1bf40b7d-3f26-4235-aa97-875a41f0e4fc/scratchpad/"
out={}
for mode in (True,False):
    wb=openpyxl.load_workbook(P, read_only=True, data_only=mode)
    ws=wb["Dropdowns"]
    cols=collections.defaultdict(list)
    for i,row in enumerate(ws.iter_rows(values_only=True), start=1):
        for j,v in enumerate(row,1):
            if v is not None and str(v).strip()!="":
                cols[j].append((i,v))
    out["values" if mode else "formulas"]=cols
    wb.close()
V=out["values"]; F=out["formulas"]
json.dump({gl(j):[[i,str(v)] for i,v in lst] for j,lst in V.items()}, open(SP+"bm_dropdowns.json","w"), indent=1)
print("Dropdowns sheet: populated columns:", [gl(j) for j in sorted(V)])
for j in sorted(V):
    lst=V[j]
    print(f"\n===== COL {gl(j)}  n={len(lst)}  rows {lst[0][0]}..{lst[-1][0]} =====")
    for i,v in lst[:40]:
        f=dict(F.get(j,[])).get(i)
        fs=f"   <<{f[:110]}>>" if isinstance(f,str) and f.startswith("=") else ""
        print(f"  {gl(j)}{i}: {str(v)[:130]}{fs}")
    if len(lst)>40: print(f"  ... (+{len(lst)-40} more)")
