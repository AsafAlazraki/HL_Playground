import openpyxl, re, collections
from openpyxl.utils import get_column_letter
SRC = r"C:/Users/AsafA/Downloads/MASTER PRICE FILE.xlsx"
_wv=_wf=None
def books():
    global _wv,_wf
    if _wv is None:
        _wv = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
        _wf = openpyxl.load_workbook(SRC, data_only=False)   # normal mode -> shared formulas translated
    return _wv,_wf
def grid(sheet, maxr=None, maxc=None):
    wv,wf=books(); ws=wv[sheet]; wsf=wf[sheet]
    mr=maxr or ws.max_row; mc=maxc or ws.max_column
    V=[[None]*(mc+1) for _ in range(mr+1)]
    F=[[None]*(mc+1) for _ in range(mr+1)]
    for i,row in enumerate(ws.iter_rows(min_row=1,max_row=mr,max_col=mc,values_only=True),1):
        for j,v in enumerate(row,1): V[i][j]=v
    for i,row in enumerate(wsf.iter_rows(min_row=1,max_row=mr,max_col=mc,values_only=True),1):
        for j,v in enumerate(row,1): F[i][j]=v
    return V,F,mr,mc
def norm(f, r):
    """normalise row numbers: =D5*E5 with r=5 -> =D{r}*E{r}"""
    if not isinstance(f,str): return f
    return re.sub(r'(?<![A-Za-z0-9_$])(\$?[A-Z]{1,3})\$?%d(?![0-9])'%r, r'\1{r}', f)
def cellstr(L,i,v,f,flen=200,vlen=90):
    s=f"{L}{i}="
    if isinstance(f,str) and f.startswith("="):
        s+="[F]"+f[:flen]
        if v is not None: s+=" ->"+repr(v)[:vlen]
    else: s+=repr(v)[:vlen]
    return s
